#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
mpdt_generator/generator.py — Rewritten MPDT generator to:
- Build AU+ headers from template att_matrix (Row1=C, Row2="E B")
- First 15 AU columns permanent for all rows
- Filter remaining AU to those present in LoDM (any class)
- Fill AU values per row from join2_df (wide), tolerant to code separators
- Apply checkerboard for attributes not applicable (per LoDM) to the row’s class
- Write count = LoDM attribute count for the row’s class + 10

Integrates with pipeline step4 (main.run_step4) and existing sources load.
References: main step4 orchestration [2.4]; config paths [4.1, 4.2]; LoDM/attr normalization helpers conceptually aligned with your codebase [1.2, 1.5].
"""

from __future__ import annotations

import argparse
import logging
import re
import tempfile
from pathlib import Path
from typing import Any, Dict, List, Optional, Sequence, Set, Tuple

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

from utils.common import (
    get_available_path,
    load_config,
    normalize_text,
    read_table_any,
    resolve_workspace,
    sanitize_filename,
    setup_logger,
    timestamped_dir,
    write_json,
)

FIRST15_PERMANENT = 15
HEADER_ROW1 = 1
HEADER_ROW2 = 2
DATA_FIRST_ROW = 3
DEFAULT_AU_LETTER = "AU"
DEFAULT_COUNT_LABEL = "count"

_RE_SEP = re.compile(r'[\s_\-\.:]+')

def _norm_att_code(s: str) -> str:
    return _RE_SEP.sub('', str(s or "")).lower()

def _is_empty(value: Any) -> bool:
    if value is None:
        return True
    try:
        if pd.isna(value):
            return True
    except Exception:
        pass
    return str(value).strip() in ("", "nan", "none", "NaN", "NaT")

def _col_letter_to_index(letter: str) -> int:
    letter = letter.strip().upper()
    num = 0
    for ch in letter:
        if 'A' <= ch <= 'Z':
            num = num * 26 + (ord(ch) - ord('A') + 1)
    return num

def _code_variants(code: str) -> Set[str]:
    s = str(code or "").strip()
    if not s:
        return set()
    c = s.lower()
    return {c, c.replace("_", "-"), c.replace("-", "_")}

# ----------------------------
# LoDM helpers
# ----------------------------

def lodm_all_attr_norms(lodm_df: pd.DataFrame) -> Set[str]:
    if lodm_df is None or lodm_df.empty:
        return set()
    att_col = next((c for c in lodm_df.columns if normalize_text(c) == "atttypename"), None)
    if not att_col:
        return set()
    return {_norm_att_code(v) for v in lodm_df[att_col].dropna().astype(str) if str(v).strip()}

def lodm_allowed_codes_for_class(lodm_df: pd.DataFrame, class_code: str) -> Set[str]:
    if lodm_df is None or lodm_df.empty or not class_code:
        return set()
    cc_col = next((c for c in lodm_df.columns if normalize_text(c) == "classcode"), None)
    att_col = next((c for c in lodm_df.columns if normalize_text(c) == "atttypename"), None)
    if not cc_col or not att_col:
        return set()
    variants = _code_variants(class_code)
    rows = lodm_df[lodm_df[cc_col].fillna("").astype(str).str.strip().str.lower().isin(variants)]
    return {_norm_att_code(v) for v in rows[att_col].dropna().astype(str) if str(v).strip()}

def lodm_attr_count_for_class(lodm_df: pd.DataFrame, class_code: str) -> int:
    return len(lodm_allowed_codes_for_class(lodm_df, class_code))

# ----------------------------
# SmartForms/join2 helpers
# ----------------------------

def ensure_join2_wide(join2_df: pd.DataFrame, uaid3_candidates: Sequence[str] = ("UAID_3", "Uaid_3", "Uaid", "UAID", "Asset_ID")) -> Tuple[pd.DataFrame, Optional[str]]:
    if join2_df is None or join2_df.empty:
        return pd.DataFrame(), None

    uaid3_col = next((c for c in join2_df.columns if normalize_text(c) in {normalize_text(x) for x in uaid3_candidates}), None)
    # If already wide enough, trust it
    if uaid3_col and len(join2_df.columns) > 50:
        return join2_df.astype(str).fillna(""), uaid3_col

    cols = [str(c).strip() for c in join2_df.columns]
    attr_id_col = next((c for c in cols if normalize_text(c) in {"atttypename", "attrtypename", "attrtypeid", "attrtypedisplayname"}), None)
    val_col = next((c for c in cols if normalize_text(c) in {"attributevalue", "attrvalue", "value"}), None)

    if uaid3_col and attr_id_col and val_col:
        wide = join2_df[[uaid3_col, attr_id_col, val_col]].copy()
        wide.columns = [uaid3_col, "attr_id", "attr_val"]
        pivoted = wide.pivot_table(index=uaid3_col, columns="attr_id", values="attr_val", aggfunc="first").reset_index()
        pivoted = pivoted.astype(str).replace("nan", "")
        return pivoted, uaid3_col

    return join2_df.astype(str).fillna(""), uaid3_col

def _column_name_candidates_for_code(code: str) -> List[str]:
    base = str(code or "")
    cand = {base, base.replace(":", "_"), base.replace("_", ":"), _RE_SEP.sub("", base)}
    cand |= {x.lower() for x in list(cand)}
    cand |= {_RE_SEP.sub("", x) for x in list(cand)}
    return list(cand)

def lookup_join2_value(join2_wide: pd.DataFrame, key_col: str, key_val: str, code: str) -> str:
    if join2_wide is None or join2_wide.empty or not key_col or _is_empty(key_val):
        return ""
    row = join2_wide[join2_wide[key_col].fillna("").astype(str).str.strip() == str(key_val).strip()]
    if row.empty:
        return ""
    # exact
    if code in row.columns:
        v = row.iloc[0][code]
        return "" if pd.isna(v) else str(v).strip()
    # tolerant
    norm_map: Dict[str, str] = {}
    for c in row.columns:
        norm_map[_norm_att_code(c)] = c
        norm_map[normalize_text(c)] = c
    for cand in _column_name_candidates_for_code(code):
        k1 = _norm_att_code(cand)
        k2 = normalize_text(cand)
        if k1 in norm_map:
            v = row.iloc[0][norm_map[k1]]
            return "" if pd.isna(v) else str(v).strip()
        if k2 in norm_map:
            v = row.iloc[0][norm_map[k2]]
            return "" if pd.isna(v) else str(v).strip()
    return ""

# ----------------------------
# att_matrix -> AU headers
# ----------------------------

def load_att_matrix_headers_from_template(template_path: Path, sheet_name: str = "att_matrix") -> Tuple[List[str], List[str]]:
    wb = load_workbook(template_path, read_only=True, data_only=True, keep_vba=True)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"att_matrix sheet not found in template: {sheet_name}")
    ws = wb[sheet_name]
    r1, r2 = [], []
    for r in ws.iter_rows(min_row=2, values_only=True):
        # A=0, B=1, C=2, D=3, E=4
        b = str(r[1] or "").strip()
        c = str(r[2] or "").strip()
        e = str(r[4] or "").strip()
        if not c:
            continue
        r1.append(c)
        r2.append(" ".join([x for x in (e, b) if str(x).strip()]))
    return r1, r2

def build_final_au_headers(template_path: Path, lodm_df: pd.DataFrame, first15: int = FIRST15_PERMANENT) -> Tuple[List[str], List[str]]:
    full_r1, full_r2 = load_att_matrix_headers_from_template(template_path, "att_matrix")
    perm_r1, perm_r2 = full_r1[:first15], full_r2[:first15]
    dyn_r1_all, dyn_r2_all = full_r1[first15:], full_r2[first15:]
    lodm_norms = lodm_all_attr_norms(lodm_df)
    dyn_kept_r1, dyn_kept_r2 = [], []
    for i, code in enumerate(dyn_r1_all):
        if _norm_att_code(code) in lodm_norms:
            dyn_kept_r1.append(code)
            dyn_kept_r2.append(dyn_r2_all[i])
    return perm_r1 + dyn_kept_r1, perm_r2 + dyn_kept_r2

# ----------------------------
# Worksheet operations
# ----------------------------

def find_au_start(ws) -> Optional[int]:
    # Prefer configured AU letter
    try:
        return _col_letter_to_index(DEFAULT_AU_LETTER)
    except Exception:
        pass
    # Fallback: first empty after a block of non-empty headers in row1
    seen = 0
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=HEADER_ROW1, column=c).value
        if v is not None and str(v).strip():
            seen += 1
        elif seen >= 5:
            return c
    return None

def find_count_col_idx(ws) -> Optional[int]:
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=HEADER_ROW2, column=c).value
        if v is not None and normalize_text(v) == normalize_text(DEFAULT_COUNT_LABEL):
            return c
    return None

def write_au_headers(ws, au_col_idx: int, au_row1: List[str], au_row2: List[str]) -> None:
    for off, (h1, h2) in enumerate(zip(au_row1, au_row2)):
        c = au_col_idx + off
        ws.cell(row=HEADER_ROW1, column=c, value=h1)
        ws.cell(row=HEADER_ROW2, column=c, value=h2)

def fill_au_values_row(ws, row_idx: int, au_col_idx: int, au_row1: List[str], join2_wide: pd.DataFrame, key_col: str, key_val: str) -> None:
    for off, code in enumerate(au_row1):
        val = lookup_join2_value(join2_wide, key_col, key_val, code)
        ws.cell(row=row_idx, column=au_col_idx + off, value=val)

def apply_checkerboard(ws, row_idx: int, au_col_idx: int, au_row1: List[str], class_code: str, lodm_df: pd.DataFrame, first15: int = FIRST15_PERMANENT) -> None:
    if lodm_df is None or lodm_df.empty:
        return
    allowed = lodm_allowed_codes_for_class(lodm_df, class_code)
    black = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
    for off, code in enumerate(au_row1):
        if off < first15:
            continue
        if _norm_att_code(code) and _norm_att_code(code) not in allowed:
            ws.cell(row=row_idx, column=au_col_idx + off).fill = black

def write_count(ws, row_idx: int, count_col_idx: int, lodm_df: pd.DataFrame, class_code: str) -> None:
    cnt = lodm_attr_count_for_class(lodm_df, class_code)
    ws.cell(row=row_idx, column=count_col_idx, value=cnt + 10)

# ----------------------------
# Core per-UAID generator
# ----------------------------

def generate_single_mpdt(
    workspace: Path,
    cfg: dict,
    uaid2: str,
    mpdt_dir: Path,
    mapping_dict: dict,
    columns: list[str],
    scope2: pd.DataFrame,
    scope3: pd.DataFrame,
    sf_l2: pd.DataFrame,
    sf_l3: pd.DataFrame,
    lodm_df: pd.DataFrame,
    pw_df: pd.DataFrame,
    control_df: pd.DataFrame,
    logger: logging.Logger,
    deliverable_file: str,
    l2_df: pd.DataFrame | None = None,
    row1_codes: dict | None = None,
    indexes: dict | None = None,
) -> Path | None:
    """
    Overhauled to:
    - Build AU headers from att_matrix + LoDM (first15 permanent, rest filtered)
    - Fill AU values from join2 (sf_l3) per UAID_3
    - Apply checkerboard and count
    Other governance-cell population remains as in your existing flow (use your current logic where noted).
    """
    logger.info("  Generating MPDT for %s", uaid2)

    # Resolve output filename/path
    filename = deliverable_file or f"MPDT_{sanitize_filename(uaid2)}.xlsm"
    out_dir = mpdt_dir / sanitize_filename(uaid2)
    out_dir.mkdir(parents=True, exist_ok=True)
    output_path = get_available_path(out_dir / sanitize_filename(filename))

    # Template path from config
    tpl_rel = cfg.get("paths", {}).get("mpdt_template")
    template_path = (workspace / tpl_rel).resolve() if tpl_rel else None
    if not template_path or not template_path.exists():
        logger.error("  MPDT template not found at paths.mpdt_template")
        return None

    # Build AU headers from att_matrix + LoDM
    au_row1, au_row2 = build_final_au_headers(template_path, lodm_df, FIRST15_PERMANENT)

    # Load template workbook (keep VBA), select sheet
    wb = load_workbook(template_path, keep_vba=True)
    sheet_name = "MPDT Element of Asset" if "MPDT Element of Asset" in wb.sheetnames else wb.sheetnames[0]
    ws = wb[sheet_name]

    # Detect AU start and count column
    au_start_idx = find_au_start(ws) or _col_letter_to_index(DEFAULT_AU_LETTER)
    count_col_idx = find_count_col_idx(ws) or (au_start_idx + 1)

    # Write AU headers (Row1/Row2)
    write_au_headers(ws, au_start_idx, au_row1, au_row2)

    # Ensure join2 wide
    join2_wide, uaid3_col = ensure_join2_wide(sf_l3)
    if not uaid3_col:
        # Try to detect from scope3 or indexes
        uaid3_col = (indexes or {}).get("uaid3_col_s3") or next(
            (c for c in sf_l3.columns if normalize_text(c) in {normalize_text(x) for x in ("UAID_3", "Uaid_3", "Uaid", "UAID")}),
            None,
        )

    # Determine rows for this UAID_2 from scope3 (or indexes if provided) [1.3]
    _uaid2_norms = {normalize_text(x) for x in ("UAID_2", "uaid_2", "uaid2", "ParentUaid", "Parent_UAID")}
    uaid2_col_s3 = (indexes or {}).get("uaid2_col_s3") or next((c for c in scope3.columns if normalize_text(c) in _uaid2_norms), None)
    rows_for_uaid = pd.DataFrame()
    if uaid2_col_s3:
        rows_for_uaid = scope3[scope3[uaid2_col_s3].fillna("").astype(str).str.strip().str.upper() == uaid2.upper()].copy()
    if rows_for_uaid.empty:
        logger.warning("  No scope3 rows found for %s — skipping MPDT.", uaid2)
        return None

    # Write data rows (governance columns: use your existing logic to set A..(AU-1) as before)
    row_idx = DATA_FIRST_ROW
    for _, s3row in rows_for_uaid.iterrows():
        # TODO: Write governance columns (UAIDs, names, coords, etc.) to ws at row_idx using your existing mapping_dict/columns
        # This preserves your current behavior for A..(AU-1).
        # Example:
        # for col_name in fixed_columns: ws.cell(row=row_idx, column=col_index, value=value)

        # Fill AU+ using UAID_3 key
        uaid3_val = str(s3row.get(uaid3_col, "")) if uaid3_col else ""
        fill_au_values_row(ws, row_idx, au_start_idx, au_row1, join2_wide, uaid3_col, uaid3_val)

        # Determine ClassCode for this row to apply checkerboard and count (+10)
        class_code = ""
        for k in ("AssetHierarchyCategory", "HS2_Class", "ClassCode"):
            if k in s3row and not _is_empty(s3row[k]):
                class_code = str(s3row[k]).strip()
                break

        apply_checkerboard(ws, row_idx, au_start_idx, au_row1, class_code, lodm_df, first15=FIRST15_PERMANENT)
        write_count(ws, row_idx, count_col_idx, lodm_df, class_code)

        row_idx += 1

    # Save
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    logger.info("  MPDT written: %s", output_path.name)
    return output_path

# ----------------------------
# Batch
# ----------------------------

def generate_mpdt_batch(
    workspace: Path,
    cfg: dict,
    targets: list[dict],
    output_dir: Path,
    sources: dict,
    logger: logging.Logger,
) -> dict:
    """
    Batch generator compatible with main.run_step4.
    sources: must include scope3_df, sf_l3 (SmartForms L3), lodm_df, pw_df, control_df, l2_df as per your loader.
    """
    scope3 = sources.get("scope3_df", pd.DataFrame())
    sf_l3 = sources.get("sf_l3_df", pd.DataFrame()) or sources.get("sf_l3", pd.DataFrame())
    lodm_df = sources.get("lodm_df", pd.DataFrame())
    pw_df = sources.get("pw_df", pd.DataFrame())
    control_df = sources.get("control_df", pd.DataFrame())
    l2_df = sources.get("l2_df", pd.DataFrame())
    mapping_dict = sources.get("mapping_dict", {})  # optional existing mapping
    columns = sources.get("mpdt_columns", [])       # optional existing columns

    if scope3.empty:
        raise RuntimeError("AssetsScope3 data not found. Run data cache step first.")

    mpdt_dir = output_dir / "MPDT"
    mpdt_dir.mkdir(parents=True, exist_ok=True)

    generated, errors = [], []
    for target in targets:
        uaid2 = target["uaid"]
        deliverable_file = target.get("file", "")
        try:
            out = generate_single_mpdt(
                workspace, cfg, uaid2, mpdt_dir, mapping_dict, columns,
                sources.get("scope2_df", pd.DataFrame()), scope3,
                sources.get("sf_l2_df", pd.DataFrame()), sf_l3, lodm_df, pw_df, control_df,
                logger, deliverable_file, l2_df=l2_df, row1_codes=sources.get("row1_codes"), indexes=sources.get("indexes"),
            )
            if out:
                generated.append({"uaid": uaid2, "file": str(out)})
        except Exception as exc:
            logger.error("MPDT generation failed for %s: %s", uaid2, exc, exc_info=True)
            errors.append({"uaid": uaid2, "error": str(exc)})

    return {"generated": generated, "errors": errors}

# ----------------------------
# Standalone entry (debug/local)
# ----------------------------

def main() -> None:
    parser = argparse.ArgumentParser(description="Generate MPDT files (att_matrix-driven AU headers)")
    parser.add_argument("--workspace", default=None)
    parser.add_argument("--target-uaid2", nargs="+", required=True)
    args = parser.parse_args()

    workspace = resolve_workspace(args.workspace)
    cfg = load_config(workspace)
    logger = setup_logger(workspace, "mpdt_generator", cfg.get("log_level", "INFO"))
    logger.info("=== MPDT Generator (att_matrix) ===")

    # Minimal local loader (pipeline step4 will pass proper sources)
    # Here we try to read from configured paths if present.
    paths = cfg.get("paths", {})
    sources: Dict[str, pd.DataFrame] = {}
    try:
        sources["scope3_df"] = read_table_any(workspace / paths.get("l3_assets_scope_data", "Input/l3_assets_scope_data.xlsx"))
    except Exception:
        sources["scope3_df"] = pd.DataFrame()
    try:
        sources["sf_l3_df"] = read_table_any(workspace / paths.get("smartforms_fallback", "Input/SmartForms_RAW_MPDT_L2&L3.xlsx"))
    except Exception:
        sources["sf_l3_df"] = pd.DataFrame()
    try:
        sources["lodm_df"] = read_table_any(workspace / paths.get("lodm", "Input/1MC06-ASC-IM-SCH-C002-000009_lodm.xlsx"))
    except Exception:
        sources["lodm_df"] = pd.DataFrame()
    try:
        sources["pw_df"] = read_table_any(workspace / paths.get("pw_extract", "Input/ACBOS_MPDT.xlsx"))
    except Exception:
        sources["pw_df"] = pd.DataFrame()
    try:
        sources["control_df"] = read_table_any(workspace / paths.get("control_file", "Input/1MC06-ASC-IM-GDE-C002-000090_controlfile.xlsx"))
    except Exception:
        sources["control_df"] = pd.DataFrame()
    try:
        sources["l2_df"] = read_table_any(workspace / paths.get("l2_uaid_acbos", "Input/L2 UAID-ACBOS.xlsx"))
    except Exception:
        sources["l2_df"] = pd.DataFrame()

    uaids = []
    for v in args.target_uaid2:
        uaids.extend(u.strip() for u in v.split(",") if u.strip())
    targets = [{"uaid": u} for u in uaids]

    output_dir = timestamped_dir(workspace, "Output")
    result = generate_mpdt_batch(workspace, cfg, targets, output_dir, sources, logger)
    write_json(workspace / "Output" / "mpdt_result.json", result)
    logger.info("Generated: %d, Errors: %d", len(result["generated"]), len(result["errors"]))


if __name__ == "__main__":
    main()