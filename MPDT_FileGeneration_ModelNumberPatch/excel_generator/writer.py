"""
excel_generator/writer.py — Write pipeline outputs to Excel.
"""
from __future__ import annotations

import logging
import os
import zipfile
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.comments import Comment


def write_excel(
    df: pd.DataFrame,
    output_path: Path,
    sheet_name: str,
    logger: logging.Logger,
) -> None:
    """Write a DataFrame to Excel with ZIP integrity check."""
    output_path.parent.mkdir(parents=True, exist_ok=True)
    tmp_path = output_path.with_suffix(".tmp.xlsx")

    with pd.ExcelWriter(tmp_path, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name=sheet_name, index=False)

    with zipfile.ZipFile(tmp_path, "r") as zf:
        bad = zf.testzip()
        if bad is not None:
            raise RuntimeError(f"Generated workbook is corrupt at entry: {bad}")

    os.replace(tmp_path, output_path)
    logger.info("Wrote: %s (%d rows)", output_path.name, len(df))


def write_multi_sheet_excel(
    dfs: dict[str, pd.DataFrame],
    output_path: Path,
    logger: logging.Logger,
) -> None:
    """Write multiple DataFrames to separate sheets in one workbook."""
    output_path.parent.mkdir(parents=True, exist_ok=True)
    tmp_path = output_path.with_suffix(".tmp.xlsx")

    with pd.ExcelWriter(tmp_path, engine="openpyxl") as writer:
        for sheet_name, df in dfs.items():
            df.to_excel(writer, sheet_name=sheet_name, index=False)

    with zipfile.ZipFile(tmp_path, "r") as zf:
        bad = zf.testzip()
        if bad is not None:
            raise RuntimeError(f"Generated workbook is corrupt at entry: {bad}")

    os.replace(tmp_path, output_path)
    logger.info("Wrote: %s (%d sheets)", output_path.name, len(dfs))


def write_classification_output(
    classification_result: dict,
    output_path: Path,
    logger: logging.Logger,
) -> None:
    """Write classification results to Excel with comments for skipped UAIDs."""
    records = classification_result.get("csv_records", [])
    if not records:
        logger.warning("No classification records to write.")
        return

    df = pd.DataFrame(records)
    dfs = {"Classification": df}

    # Add conflict/skipped sheets if any
    conflicts = classification_result.get("conflicts", [])
    if conflicts:
        dfs["Conflicts"] = pd.DataFrame(conflicts)

    skipped = classification_result.get("skipped", [])
    if skipped:
        dfs["Needs_Review"] = pd.DataFrame(skipped)

    write_multi_sheet_excel(dfs, output_path, logger)

    # Add comments to skipped UAIDs in Classification sheet
    if skipped:
        try:
            wb = load_workbook(output_path)
            ws = wb["Classification"]
            
            # Find the row for each skipped UAID and add comment
            for skipped_item in skipped:
                skipped_uaid = skipped_item.get("uaid", "")
                skip_reason = skipped_item.get("notes", "")
                
                # Find the row with this UAID (column A, starting from row 2 after header)
                for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
                    if row[0].value == skipped_uaid:
                        # Add comment to the UAID cell
                        cell = row[0]
                        cell.comment = Comment(
                            f"SKIPPED: {skip_reason}\n\nGeneration output not created for this asset.",
                            "Pipeline"
                        )
                        break
            
            wb.save(output_path)
            logger.info("Added skip comments to classification output")
        except Exception as e:
            logger.warning("Could not add comments to classification output: %s", e)


def add_skip_comments_to_asset_deliverables(
    deliverables_path: Path,
    skipped_uaids: list[dict],
    logger: logging.Logger,
) -> None:
    """Add comments to asset_deliverables.xlsx for skipped target UAIDs."""
    if not skipped_uaids or not deliverables_path.exists():
        return
    
    try:
        wb = load_workbook(deliverables_path)
        ws = wb.active
        
        # Build a set of skipped UAIDs with their skip reasons
        skip_map = {item.get("uaid", ""): item.get("notes", "") for item in skipped_uaids}
        
        # Find ASSET_ID column (case-insensitive)
        asset_id_col = None
        for col_idx, cell in enumerate(ws[1], start=1):
            if cell.value and str(cell.value).upper() == "ASSET_ID":
                asset_id_col = col_idx
                break
        
        if asset_id_col is None:
            logger.warning("Could not find ASSET_ID column in asset_deliverables")
            return
        
        # Add comments to rows matching skipped UAIDs
        for row_idx in range(2, ws.max_row + 1):
            asset_id = ws.cell(row_idx, asset_id_col).value
            if asset_id in skip_map:
                cell = ws.cell(row_idx, asset_id_col)
                cell.comment = Comment(
                    f"SKIPPED: {skip_map[asset_id]}\n\nNo generation output created for this asset.",
                    "Pipeline"
                )
        
        wb.save(deliverables_path)
        logger.info("Added skip comments to asset_deliverables.xlsx")
    except Exception as e:
        logger.warning("Could not add comments to asset_deliverables: %s", e)


def add_mapping_method_to_asset_deliverables(
    deliverables_path: Path,
    classification_records: list[dict],
    logger: logging.Logger,
) -> None:
    """Add/refresh Deliverable_Mapped_By column in asset_deliverables.xlsx from classification records."""
    if not deliverables_path.exists() or not classification_records:
        return

    try:
        wb = load_workbook(deliverables_path)
        ws = wb.active

        # Find required columns.
        headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
        asset_col = None
        pw_uaid_col = None
        mapped_col = None
        for idx, h in enumerate(headers, start=1):
            hs = str(h).strip().upper() if h is not None else ""
            if hs == "ASSET_ID":
                asset_col = idx
            if hs == "PW_UAID":
                pw_uaid_col = idx
            if hs == "DELIVERABLE_MAPPED_BY":
                mapped_col = idx

        if asset_col is None:
            logger.warning("Could not find ASSET_ID column in asset_deliverables")
            return

        if mapped_col is None:
            mapped_col = ws.max_column + 1
            ws.cell(1, mapped_col).value = "Deliverable_Mapped_By"

        map_by_uaid = {
            str(r.get("uaid", "")).strip(): str(r.get("deliverable_mapped_by", "")).strip()
            for r in classification_records
            if str(r.get("uaid", "")).strip()
        }

        for r in range(2, ws.max_row + 1):
            uaid_asset = str(ws.cell(r, asset_col).value or "").strip()
            uaid_pw = str(ws.cell(r, pw_uaid_col).value or "").strip() if pw_uaid_col else ""
            mapped = map_by_uaid.get(uaid_asset) or map_by_uaid.get(uaid_pw) or ""
            ws.cell(r, mapped_col).value = mapped

        wb.save(deliverables_path)
        logger.info("Added Deliverable_Mapped_By column to asset_deliverables.xlsx")
    except Exception as e:
        logger.warning("Could not add Deliverable_Mapped_By to asset_deliverables: %s", e)
