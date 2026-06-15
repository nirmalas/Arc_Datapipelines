# generator_patch_attmatrix.py
import re
import pandas as pd
from openpyxl import load_workbook

def _norm_text(s: str) -> str:
    return re.sub(r"\s+", " ", str(s or "").strip()).lower()

def load_att_matrix_headers(template_path: str, sheet_name: str = "att_matrix") -> tuple[list[str], list[str]]:
    """
    Return (row1_labels, row2_labels) for AU+ columns derived from the template's att_matrix:
    - Row 1: column C values
    - Row 2: "E B" (E and B columns concatenated with a space; strip empties)
    Rows with empty column C are ignored.
    """
    wb = load_workbook(template_path, read_only=True, data_only=True, keep_vba=True)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"att_matrix sheet not found in template: {sheet_name}")
    ws = wb[sheet_name]
    row1: list[str] = []
    row2: list[str] = []
    # Assuming headers row is 1 and data starts at row 2; adjust if needed
    for r in ws.iter_rows(min_row=2, values_only=True):
        col_b = str(r[1] or "").strip()  # B
        col_c = str(r[2] or "").strip()  # C
        col_e = str(r[4] or "").strip()  # E
        if not col_c:
            continue
        r1 = col_c
        # Combine E and B, skipping empties and duplicates
        parts = [p for p in [col_e, col_b] if p]
        r2 = " ".join(parts).strip()
        row1.append(r1)
        row2.append(r2)
    return row1, row2

def build_au_headers_from_att_matrix(template_path: str, first15_permanent: int = 15) -> dict:
    """
    Returns:
      {
        'row1_all': [ ... AU+ all from att_matrix ... ],
        'row2_all': [ ... ],
        'row1_perm': [ first 15 row1 values ],
        'row2_perm': [ first 15 row2 values ],
        'row1_dynamic': [ remainder ],
        'row2_dynamic': [ remainder ],
    }
    """
    r1_all, r2_all = load_att_matrix_headers(template_path, "att_matrix")
    row1_perm = r1_all[:first15_permanent]
    row2_perm = r2_all[:first15_permanent]
    row1_dyn = r1_all[first15_permanent:]
    row2_dyn = r2_all[first15_permanent:]
    return {
        "row1_all": r1_all,
        "row2_all": r2_all,
        "row1_perm": row1_perm,
        "row2_perm": row2_perm,
        "row1_dynamic": row1_dyn,
        "row2_dynamic": row2_dyn,
    }