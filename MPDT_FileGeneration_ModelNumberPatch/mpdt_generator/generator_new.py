#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
mpdt_generator/generator.py — MPDT generator:
- AU Row1 from template att_matrix col C (short code)
- AU Row2 from template att_matrix ("E B")
- First 15 AU columns permanent
- Remaining AU columns filtered to LoDM attributes
- Fill AU+ values from SmartForms L3 (wide) keyed by UAID_3
- Checkerboard: blacken non-applicable attributes per LoDM class (beyond first 15)
- Count = LoDM attribute count + 10

Integrates with main.py step4 (run_step4 → generate_mpdt_batch).

Assumptions:
- sources["sf_l3_df"] (or "sf_l3") is a wide table with UAID_3 and attribute columns like Com_Dscrptn, CM_Chg, CM_CRT, NetVolume, etc. (from your sample).
- sources["scope3_df"] has UAID_2 and UAID_3 rows to iterate.
"""

from __future__ import annotations

import argparse
import logging
import re
from pathlib import Path
from typing import Any, Dict, List, Optional, Sequence, Set, Tuple

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

from utils.common import (
    get_available_path,
    load_config,
    normalize_text,
    read_table_any,
    resolve_workspace,
    sanitize_filename,
    setup_logger,
    timestamped_dir,
    write_json,
)

# ----------------------------
# Constants
# ----------------------------

FIRST15_PERMANENT = 15
HEADER_ROW1 = 1
HEADER_ROW2 = 2
DATA_FIRST_ROW = 3
DEFAULT_AU_LETTER = "AU"
DEFAULT_COUNT_LABEL = "count"

_RE_SEP = re.compile(r'[\s_\-\.:]+')


# ----------------------------
# Small utilities
# ----------------------------

def _is_empty(value: Any) -> bool:
    if value is None:
        return True
    try:
        if pd.isna(value):
            return True
    except Exception:
        pass
    return str(value).strip() in ("", "nan", "none", "NaN", "NaT", "nat")

def _norm_att_code(s: str) -> str:
    return _RE_SEP.sub('', str(s or "")).lower()

def _code_variants(code: str) -> Set[str]:
    s = str(code or "").strip()
    if not s:
        return set()
    c = s.lower()
    return {c, c.replace("_", "-"), c.replace("-", "_")}

def _col_letter_to_index(letter: str) -> int:
    letter = letter.strip().upper()
    num = 0
    for ch in letter:
        if 'A' <= ch <= 'Z':
            num = num * 26 + (ord(ch) - ord('A') + 1)
    return num


# ----------------------------
# LoDM helpers
# ----------------------------

def lodm_all_attr_norms(lodm_df: pd.DataFrame) -> Set[str]:
    if lodm_df is None or lodm_df.empty:
        return set()
    att_col = next((c for c in lodm_df.columns if normalize_text(c) == "atttypename"), None)
    if not att_col:
        return set()
    return {_norm_att_code(v) for v in lodm_df[att_col].dropna().astype(str) if str(v).strip()}

def lodm_allowed_codes_for_class(lodm_df: pd.DataFrame, class_code: str) -> Set[str]:
    if lodm_df is None or lodm_df.empty or not class_code:
        return set()
    cc_col = next((c for c in lodm_df.columns if normalize_text(c) == "classcode"), None)
    att_col = next((c for c in lodm_df.columns if normalize_text(c) == "atttypename"), None)
    if not cc_col or not att_col:
        return set()
    variants = _code_variants(class_code)
    rows = lodm_df[lodm_df[cc_col].fillna("").astype(str).str.strip().str.lower().isin(variants)]
    return {_norm_att_code(v) for v in rows[att_col].dropna().astype(str) if str(v).strip()}

def lodm_attr_count_for_class(lodm_df: pd.DataFrame, class_code: str) -> int:
    return len(lodm_allowed_codes_for_class(lodm_df, class_code))

def build_lodm_maps(lodm_df: pd.DataFrame) -> Tuple[Dict[str, str], Dict[str, Set[str]]]:
    """
    code_to_id: normalized AttTypeName (e.g., 'comdscrptn') -> AttributeTypeId (string)
    class_to_codes: normalized ClassCode -> set of normalized codes
    """
    code_to_id: Dict[str, str] = {}
    class_to_codes: Dict[str, Set[str]] = {}
    if lodm_df is None or lodm_df.empty:
        return code_to_id, class_to_codes
    cc_col = next((c for c in lodm_df.columns if normalize_text(c) == "classcode"), None)
    name_col = next((c for c in lodm_df.columns if normalize_text(c) == "atttypename"), None)
    id_col = next((c for c in lodm_df.columns if normalize_text(c) in {"attributetypeid", "attrtypeid", "atttypeid"}), None)
    for _, r in lodm_df.iterrows():
        code_norm = _norm_att_code(r.get(name_col, ""))
        if code_norm:
            if id_col:
                att_id = str(r.get(id_col, "")).strip()
                if att_id:
                    code_to_id.setdefault(code_norm, att_id)
            if cc_col:
                cc = str(r.get(cc_col, "")).strip().lower()
                if cc:
                    class_to_codes.setdefault(cc, set()).add(code_norm)
    return code_to_id, class_to_codes


# ----------------------------
# SmartForms L3 (join2) helpers
# ----------------------------

def _column_name_candidates_for_code(code: str) -> List[str]:
    base = str(code or "")
    cand = {base, base.replace(":", "_"), base.replace("_", ":"), _RE_SEP.sub("", base)}
    cand |= {x.lower() for x in list(cand)}
    cand |= {_RE_SEP.sub("", x) for x in list(cand)}
    return list(cand)

def ensure_join2_wide(join2_df: pd.DataFrame, logger: Optional[logging.Logger] = None) -> Tuple[pd.DataFrame, Optional[str]]:
    """
    Use SmartForms L3 as a wide table (your sample) with UAID_3 present and attribute columns already wide.
    If it's not wide (i.e., has AttributeTypeId/AttributeValue rows), we will NOT pivot here because your
    loader already supplies a wide table per your sample; but we still try to be robust.

    Returns (wide_df, uaid3_col).
    """
    if join2_df is None or join2_df.empty:
        return pd.DataFrame(), None

    # Detect UAID_3 column
    uaid3_col = next((c for c in join2_df.columns if normalize_text(c) in {
        "uaid3", "uaid_3", "uaid", "asset_id", "assetid"
    }), None)

    # Heuristic: if we have > 100 columns and UAID_3 exists, treat as wide.
    if uaid3_col and len(join2_df.columns) > 50:
        return join2_df.astype(str).fillna(""), uaid3_col

    # Try pivot if normalized
    cols = [str(c).strip() for c in join2_df.columns]
    attname_col = next((c for c in cols if normalize_text(c) in {"atttypename", "attrtypename", "attrtypedisplayname"}), None)
    attid_col   = next((c for c in cols if normalize_text(c) in {"attributetypeid", "attrtypeid", "atttypeid"}), None)
    val_col     = next((c for c in cols if normalize_text(c) in {"attributevalue", "attrvalue", "value"}), None)

    if uaid3_col and val_col and (attname_col or attid_col):
        df = join2_df[[uaid3_col, attname_col or attid_col, val_col]].copy()
        df.columns = [uaid3_col, "attr_key", "attr_val"]
        piv = df.pivot_table(index=uaid3_col, columns="attr_key", values="attr_val", aggfunc="first").reset_index()
        piv = piv.astype(str).replace("nan", "")
        if logger and logger.isEnabledFor(logging.INFO):
            logger.info("SmartForms L3 pivoted to wide: rows=%d, cols=%d", len(piv), len(piv.columns))
        return piv, uaid3_col

    return join2_df.astype(str).fillna(""), uaid3_col

def lookup_join2_value(
    join2_wide: pd.DataFrame,
    key_col: str,
    key_val: str,
    au_code: str,
    code_to_id: Dict[str, str],
) -> str:
    """
    Try to read value from join2_wide row for UAID_3:
    1) AU code exact column
    2) ':'/'_' swapped, lowercase, no-separator
    3) If code maps to AttributeTypeId in LoDM, try that (string/int)
    """
    if join2_wide is None or join2_wide.empty or not key_col or _is_empty(key_val):
        return ""
    row = join2_wide[join2_wide[key_col].fillna("").astype(str).str.strip() == str(key_val).strip()]
    if row.empty:
        return ""

    # Build a normalized column map for robust matching
    norm_map: Dict[str, str] = {}
    for c in row.columns:
        norm_map[_norm_att_code(c)] = c
        norm_map[normalize_text(c)] = c

    # Direct and tolerant by code text
    for cand in _column_name_candidates_for_code(au_code):
        k1 = _norm_att_code(cand)
        k2 = normalize_text(cand)
        if k1 in norm_map:
            v = row.iloc[0][norm_map[k1]]
            return "" if pd.isna(v) else str(v).strip()
        if k2 in norm_map:
            v = row.iloc[0][norm_map[k2]]
            return "" if pd.isna(v) else str(v).strip()

    # Try mapped AttributeTypeId
    code_norm = _norm_att_code(au_code)
    att_id = code_to_id.get(code_norm)
    if att_id:
        # Try as-is, string, and int
        for candidate in (att_id, str(att_id)):
            if candidate in row.columns:
                v = row.iloc[0][candidate]
                return "" if pd.isna(v) else str(v).strip()
        if str(att_id).isdigit():
            int_id = int(str(att_id))
            if int_id in row.columns:
                v = row.iloc[0][int_id]
                return "" if pd.isna(v) else str(v).strip()

    return ""


# ----------------------------
# att_matrix → AU headers
# ----------------------------

def load_att_matrix_headers_from_template(template_path: Path, sheet_name: str = "att_matrix") -> Tuple[List[str], List[str]]:
    wb = load_workbook(template_path, read_only=True, data_only=True, keep_vba=True)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"att_matrix sheet not found in template: {sheet_name}")
    ws = wb[sheet_name]
    r1, r2 = [], []
    for r in ws.iter_rows(min_row=2, values_only=True):
        # A=0, B=1, C=2, D=3, E=4
        b = str(r[1] or "").strip()
        c = str(r[2] or "").strip()
        e = str(r[4] or "").strip()
        if not c:
            continue
        r1.append(c)
        r2.append(" ".join([x for x in (e, b) if str(x).strip()]))
    return r1, r2

def build_final_au_headers(template_path: Path, lodm_df: pd.DataFrame, first15: int = FIRST15_PERMANENT) -> Tuple[List[str], List[str]]:
    full_r1, full_r2 = load_att_matrix_headers_from_template(template_path, "att_matrix")
    perm_r1, perm_r2 = full_r1[:first15], full_r2[:first15]
    dyn_r1_all, dyn_r2_all = full_r1[first15:], full_r2[first15:]
    lodm_norms = lodm_all_attr_norms(lodm_df)
    dyn_keep_r1, dyn_keep_r2 = [], []
    for i, code in enumerate(dyn_r1_all):
        if _norm_att_code(code) in lodm_norms:
            dyn_keep_r1.append(code)
            dyn_keep_r2.append(dyn_r2_all[i])
    return perm_r1 + dyn_keep_r1, perm_r2 + dyn_keep_r2


# ----------------------------
# Worksheet operations
# ----------------------------

def find_au_start(ws) -> Optional[int]:
    try:
        return _col_letter_to_index(DEFAULT_AU_LETTER)
    except Exception:
        pass
    # Fallback scan row1
    seen = 0
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=HEADER_ROW1, column=c).value
        if v is not None and str(v).strip():
            seen += 1
        elif seen >= 5:
            return c
    return None

def find_count_col_idx(ws) -> Optional[int]:
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=HEADER_ROW2, column=c).value
        if v is not None and normalize_text(v) == normalize_text(DEFAULT_COUNT_LABEL):
            return c
    return None

def write_au_headers(ws, au_col_idx: int, au_row1: List[str], au_row2: List[str]) -> None:
    for off, (h1, h2) in enumerate(zip(au_row1, au_row2)):
        c = au_col_idx + off
        ws.cell(row=HEADER_ROW1, column=c, value=h1)
        ws.cell(row=HEADER_ROW2, column=c, value=h2)

def clear_au_region(ws, start_row: int, end_row: int, au_col_idx: int, au_len: int) -> None:
    for rr in range(start_row, end_row + 1):
        for cc in range(au_col_idx, au_col_idx + au_len):
            ws.cell(row=rr, column=cc, value=None)

def fill_au_values_row(
    ws,
    row_idx: int,
    au_col_idx: int,
    au_row1: List[str],
    join2_wide: pd.DataFrame,
    key_col: str,
    key_val: str,
    code_to_id: Dict[str, str],
) -> None:
    for off, code in enumerate(au_row1):
        val = lookup_join2_value(join2_wide, key_col, key_val, code, code_to_id)
        ws.cell(row=row_idx, column=au_col_idx + off, value=val)

def apply_checkerboard(ws, row_idx: int, au_col_idx: int, au_row1: List[str], class_code: str, lodm_df: pd.DataFrame, first15: int = FIRST15_PERMANENT) -> None:
    if lodm_df is None or lodm_df.empty:
        return
    allowed = lodm_allowed_codes_for_class(lodm_df, class_code)
    black = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
    for off, code in enumerate(au_row1):
        if off < first15:
            continue
        if _norm_att_code(code) and _norm_att_code(code) not in allowed:
            ws.cell(row=row_idx, column=au_col_idx + off).fill = black

def write_count(ws, row_idx: int, count_col_idx: int, lodm_df: pd.DataFrame, class_code: str) -> None:
    cnt = lodm_attr_count_for_class(lodm_df, class_code)
    ws.cell(row=row_idx, column=count_col_idx, value=cnt + 10)


# ----------------------------
# Core per-UAID generator
# ----------------------------

def generate_single_mpdt(
    workspace: Path,
    cfg: dict,
    uaid2: str,
    mpdt_dir: Path,
    mapping_dict: dict,
    columns: list[str],
    scope2: pd.DataFrame,
    scope3: pd.DataFrame,
    sf_l2: pd.DataFrame,
    sf_l3: pd.DataFrame,
    lodm_df: pd.DataFrame,
    pw_df: pd.DataFrame,
    control_df: pd.DataFrame,
    logger: logging.Logger,
    deliverable_file: str,
    l2_df: pd.DataFrame | None = None,
    row1_codes: dict | None = None,
    indexes: dict | None = None,
) -> Path | None:
    logger.info("  Generating MPDT for %s", uaid2)

    # Resolve output filename/path
    filename = deliverable_file or f"{sanitize_filename(uaid2)}.xlsm"
    out_dir = mpdt_dir / sanitize_filename(uaid2)
    out_dir.mkdir(parents=True, exist_ok=True)
    output_path = get_available_path(out_dir / sanitize_filename(filename))

    # Template path from config
    tpl_rel = cfg.get("paths", {}).get("mpdt_template")
    template_path = (workspace / tpl_rel).resolve() if tpl_rel else None
    if not template_path or not template_path.exists():
        logger.error("  MPDT template not found at paths.mpdt_template")
        return None

    # AU headers
    au_row1, au_row2 = build_final_au_headers(template_path, lodm_df, FIRST15_PERMANENT)

    # Load workbook/sheet
    wb = load_workbook(template_path, keep_vba=True)
    sheet_name = "MPDT Element of Asset" if "MPDT Element of Asset" in wb.sheetnames else wb.sheetnames[0]
    ws = wb[sheet_name]

    # AU start and count
    au_start_idx = find_au_start(ws) or _col_letter_to_index(DEFAULT_AU_LETTER)
    count_col_idx = find_count_col_idx(ws) or (au_start_idx + 1)

    # Write AU headers
    write_au_headers(ws, au_start_idx, au_row1, au_row2)

    # Ensure SmartForms L3 is wide and keyed by UAID_3
    join2_wide, uaid3_col = ensure_join2_wide(sf_l3, logger=logger)
    if not uaid3_col:
        # Try best-effort UAID_3 name from scope3 columns
        uaid3_col = next((c for c in ["UAID_3", "Uaid_3", "UAID", "Uaid", "Asset_ID"] if c in sf_l3.columns), "")

    # LoDM maps (for AU code → AttributeTypeId fallback)
    code_to_id, _class_to_codes = build_lodm_maps(lodm_df)

    # Find scope3 rows for this UAID_2
    _u2_norms = {normalize_text(x) for x in ("UAID_2", "uaid_2", "uaid2", "ParentUaid", "Parent_UAID")}
    uaid2_col_s3 = next((c for c in scope3.columns if normalize_text(c) in _u2_norms), None)
    if not uaid2_col_s3:
        logger.error("  UAID_2 column not found in scope3 — skipping %s", uaid2)
        return None

    rows_for_uaid = scope3[scope3[uaid2_col_s3].fillna("").astype(str).str.strip().str.upper() == uaid2.upper()]
    if rows_for_uaid.empty:
        logger.warning("  No scope3 rows for UAID_2=%s — skipping", uaid2)
        return None

    # Clear AU data region
    au_len = len(au_row1)
    start_row = DATA_FIRST_ROW
    end_row = DATA_FIRST_ROW + max(0, len(rows_for_uaid) - 1)
    clear_au_region(ws, start_row, end_row, au_start_idx, au_len)

    # Write data rows
    row_idx = DATA_FIRST_ROW
    for _, s3row in rows_for_uaid.iterrows():
        # Governance A..(AU-1): keep your existing mapping/population if you have it in pipeline.
        # This sample focuses on AU+ population. Add your fixed-field writes here if needed.

        # UAID_3 key
        key_val = str(s3row.get(uaid3_col, "")) if uaid3_col else ""

        # AU+ values from SmartForms L3 wide
        fill_au_values_row(ws, row_idx, au_start_idx, au_row1, join2_wide, uaid3_col, key_val, code_to_id)

        # Class code for checkerboard + count
        class_code = ""
        for k in ("AssetHierarchyCategory", "HS2_Class", "ClassCode"):
            if k in s3row and not _is_empty(s3row[k]):
                class_code = str(s3row[k]).strip()
                break

        apply_checkerboard(ws, row_idx, au_start_idx, au_row1, class_code, lodm_df, first15=FIRST15_PERMANENT)
        write_count(ws, row_idx, count_col_idx, lodm_df, class_code)

        row_idx += 1

    # Save
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    logger.info("  MPDT written: %s", output_path.name)
    return output_path


# ----------------------------
# Batch
# ----------------------------

def _pick_df(sources: dict, primary_key: str, fallback_key: str) -> pd.DataFrame:
    df = sources.get(primary_key, None)
    if df is None or (hasattr(df, "empty") and df.empty):
        fb = sources.get(fallback_key, None)
        return fb if fb is not None else pd.DataFrame()
    return df

def generate_mpdt_batch(
    workspace: Path,
    cfg: dict,
    targets: list[dict],
    output_dir: Path,
    sources: dict,
    logger: logging.Logger,
) -> dict:
    scope3 = sources.get("scope3_df", pd.DataFrame())
    if scope3 is None or scope3.empty:
        raise RuntimeError("AssetsScope3 data not found. Run data cache step first.")

    # IMPORTANT: use the wide SmartForms L3 (the one with UAID_3 + attribute columns)
    sf_l3 = _pick_df(sources, "sf_l3_df", "sf_l3")

    scope2 = sources.get("scope2_df", pd.DataFrame())
    sf_l2 = _pick_df(sources, "sf_l2_df", "sf_l2")
    lodm_df = sources.get("lodm_df", pd.DataFrame())
    pw_df = sources.get("pw_df", pd.DataFrame())
    control_df = sources.get("control_df", pd.DataFrame())
    l2_df = sources.get("l2_df", pd.DataFrame())
    mapping_dict = sources.get("mapping_dict", {})
    columns = sources.get("mpdt_columns", [])

    mpdt_dir = output_dir / "MPDT"
    mpdt_dir.mkdir(parents=True, exist_ok=True)

    generated, errors = [], []
    for target in targets:
        uaid2 = target["uaid"]
        deliverable_file = target.get("file", "")
        try:
            out = generate_single_mpdt(
                workspace, cfg, uaid2, mpdt_dir,
                mapping_dict, columns,
                scope2, scope3, sf_l2, sf_l3,
                lodm_df, pw_df, control_df,
                logger, deliverable_file,
                l2_df=l2_df,
                row1_codes=sources.get("row1_codes"),
                indexes=sources.get("indexes"),
            )
            if out:
                generated.append({"uaid": uaid2, "file": str(out)})
        except Exception as exc:
            logger.error("MPDT generation failed for %s: %s", uaid2, exc, exc_info=True)
            errors.append({"uaid": uaid2, "error": str(exc)})

    return {"generated": generated, "errors": errors}


# ----------------------------
# Standalone (optional debug)
# ----------------------------

def main() -> None:
    parser = argparse.ArgumentParser(description="Generate MPDT files (att_matrix-driven AU headers)")
    parser.add_argument("--workspace", default=None)
    parser.add_argument("--target-uaid2", nargs="+", required=True)
    args = parser.parse_args()

    workspace = resolve_workspace(args.workspace)
    cfg = load_config(workspace)
    logger = setup_logger(workspace, "mpdt_generator", cfg.get("log_level", "INFO"))
    logger.info("=== MPDT Generator (att_matrix) ===")

    # Minimal local load for debug; main.py will pass proper sources when running step4
    paths = cfg.get("paths", {})
    sources: Dict[str, pd.DataFrame] = {}
    try:
        sources["scope3_df"] = read_table_any(workspace / paths.get("l3_assets_scope_data", "Input/l3_assets_scope_data.xlsx"))
    except Exception:
        sources["scope3_df"] = pd.DataFrame()
    try:
        sources["sf_l3_df"] = read_table_any(workspace / paths.get("smartforms_fallback", "Input/SmartForms_RAW_MPDT_L2&L3.xlsx"))
    except Exception:
        sources["sf_l3_df"] = pd.DataFrame()
    try:
        sources["lodm_df"] = read_table_any(workspace / paths.get("lodm", "Input/1MC06-ASC-IM-SCH-C002-000009_lodm.xlsx"))
    except Exception:
        sources["lodm_df"] = pd.DataFrame()
    try:
        sources["pw_df"] = read_table_any(workspace / paths.get("pw_extract", "Input/ACBOS_MPDT.xlsx"))
    except Exception:
        sources["pw_df"] = pd.DataFrame()
    try:
        sources["control_df"] = read_table_any(workspace / paths.get("control_file", "Input/1MC06-ASC-IM-GDE-C002-000090_controlfile.xlsx"))
    except Exception:
        sources["control_df"] = pd.DataFrame()
    try:
        sources["l2_df"] = read_table_any(workspace / paths.get("l2_uaid_acbos", "Input/L2 UAID-ACBOS.xlsx"))
    except Exception:
        sources["l2_df"] = pd.DataFrame()

    uaids = []
    for v in args.target_uaid2:
        uaids.extend(u.strip() for u in v.split(",") if u.strip())
    targets = [{"uaid": u} for u in uaids]

    output_dir = timestamped_dir(workspace, "Output")
    result = generate_mpdt_batch(workspace, cfg, targets, output_dir, sources, logger)
    write_json(workspace / "Output" / "mpdt_result.json", result)
    logger.info("Generated: %d, Errors: %d", len(result["generated"]), len(result["errors"]))


if __name__ == "__main__":
    main()