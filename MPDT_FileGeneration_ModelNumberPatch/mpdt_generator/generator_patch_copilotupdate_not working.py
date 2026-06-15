#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
Optimized MPDT generator:
- Preserves governance logic (A.. before AU), builds rows from scope3 + scope2 + SmartForms L3 + mapping_dict
- AU from att_matrix (Row1=C, Row2='E B'), first 15 permanent, rest filtered by LoDM
- Fills AU values via SmartForms L3 (wide) keyed by UAID_3, tolerant matching + AttributeTypeId fallback
- Applies checkerboard applicability; Count = LoDM attr count + 10
- Performance: cached expression parsing, RowResolver, batch precomputes, robust column detection
"""

from __future__ import annotations

import argparse
import copy
import logging
import re
from dataclasses import dataclass
from functools import lru_cache
from pathlib import Path
from typing import Any, Dict, List, Optional, Sequence, Set, Tuple, Mapping

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# Project utility imports expected by your codebase
from utils.common import (
    get_available_path,
    load_config,
    normalize_text,
    read_table_any,
    resolve_workspace,
    sanitize_filename,
    setup_logger,
    timestamped_dir,
    write_json,
)

try:
    from openpyxl.utils import column_index_from_string, get_column_letter
except Exception:
    column_index_from_string = None
    def get_column_letter(idx: int) -> str:
        s = ""
        while idx:
            idx, r = divmod(idx - 1, 26)
            s = chr(r + 65) + s
        return s

# ---------------- Constants ----------------
FIRST15_PERMANENT = 15
HEADER_ROW1 = 1
HEADER_ROW2 = 2
DATA_FIRST_ROW = 3
DEFAULT_AU_LETTER = "AU"
_RE_SEP = re.compile(r'[\s_\-\.:]+')
_BLACK_FILL = PatternFill(fill_type="solid", fgColor="000000")

# ---------------- Robust normalization ----------------
def _is_empty(value: Any) -> bool:
    if value is None:
        return True
    try:
        if pd.isna(value):
            return True
    except Exception:
        pass
    s = str(value).strip().lower()
    return s == "" or s in {"nan", "none", "nat"}

def _norm_att_code(s: str) -> str:
    return _RE_SEP.sub('', str(s or "")).lower()

def _code_variants(code: str) -> Set[str]:
    s = str(code or "").strip()
    if not s:
        return set()
    c = s.lower()
    return {c, c.replace("_", "-"), c.replace("-", "_")}

def _col_letter_to_index(letter: str) -> int:
    letter = letter.strip().upper()
    num = 0
    for ch in letter:
        if 'A' <= ch <= 'Z':
            num = num * 26 + (ord(ch) - ord('A') + 1)
    return num

def _norm_key(s: str) -> str:
    return re.sub(r'[^a-z0-9]+', '', str(s or '').lower())

def _upper_key(x: Any) -> str:
    return str(x).strip().upper() if x is not None else ""

def _first_col_by_keys(df: pd.DataFrame, candidate_keys: set[str]) -> str | None:
    if df is None or df.empty:
        return None
    for col in df.columns:
        if _norm_key(col) in candidate_keys:
            return col
    return None

def _build_normalized_key(df: pd.DataFrame, col: str, out_col: str) -> None:
    if df is not None and not df.empty and col in df.columns and out_col not in df.columns:
        df[out_col] = df[col].astype(str).str.strip().str.upper()

# ---------------- LoDM helpers ----------------
def lodm_all_attr_norms(lodm_df: pd.DataFrame) -> Set[str]:
    if lodm_df is None or lodm_df.empty:
        return set()
    att_col = next((c for c in lodm_df.columns if normalize_text(c) == "atttypename"), None)
    if not att_col:
        return set()
    return {_norm_att_code(v) for v in lodm_df[att_col].dropna().astype(str) if str(v).strip()}

def lodm_allowed_codes_for_class(lodm_df: pd.DataFrame, class_code: str) -> Set[str]:
    if lodm_df is None or lodm_df.empty or not class_code:
        return set()
    cc_col = next((c for c in lodm_df.columns if normalize_text(c) == "classcode"), None)
    att_col = next((c for c in lodm_df.columns if normalize_text(c) == "atttypename"), None)
    if not cc_col or not att_col:
        return set()
    variants = _code_variants(class_code)
    rows = lodm_df[lodm_df[cc_col].fillna("").astype(str).str.strip().str.lower().isin(variants)]
    return {_norm_att_code(v) for v in rows[att_col].dropna().astype(str) if str(v).strip()}

def lodm_attr_count_for_class(lodm_df: pd.DataFrame, class_code: str) -> int:
    return len(lodm_allowed_codes_for_class(lodm_df, class_code))

def build_lodm_maps(lodm_df: pd.DataFrame) -> Tuple[Dict[str, str], Dict[str, Set[str]]]:
    code_to_id: Dict[str, str] = {}
    class_to_codes: Dict[str, Set[str]] = {}
    if lodm_df is None or lodm_df.empty:
        return code_to_id, class_to_codes
    cc_col = next((c for c in lodm_df.columns if normalize_text(c) == "classcode"), None)
    name_col = next((c for c in lodm_df.columns if normalize_text(c) == "atttypename"), None)
    id_col = next((c for c in lodm_df.columns if normalize_text(c) in {"attributetypeid", "attrtypeid", "atttypeid"}), None)
    for _, r in lodm_df.iterrows():
        code_norm = _norm_att_code(r.get(name_col, ""))
        if code_norm:
            if id_col:
                att_id = str(r.get(id_col, "")).strip()
                if att_id:
                    code_to_id.setdefault(code_norm, att_id)
            if cc_col:
                cc = str(r.get(cc_col, "")).strip().lower()
                if cc:
                    class_to_codes.setdefault(cc, set()).add(code_norm)
    return code_to_id, class_to_codes

# ---------------- att_matrix → AU headers ----------------
def load_att_matrix_headers_from_template(template_path: Path, sheet_name: str = "att_matrix") -> Tuple[List[str], List[str]]:
    wb = load_workbook(str(template_path), read_only=True, data_only=True, keep_vba=True)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"att_matrix sheet not found: {sheet_name}")
    ws = wb[sheet_name]
    r1, r2 = [], []
    for row in ws.iter_rows(min_row=2, values_only=True):
        b = str(row[1] or "").strip()
        c = str(row[2] or "").strip()
        e = str(row[4] or "").strip()
        if not c:
            continue
        r1.append(c)
        r2.append(" ".join([x for x in (e, b) if str(x).strip()]))
    return r1, r2

def build_final_au_headers(template_path: Path, lodm_df: pd.DataFrame, first15: int = FIRST15_PERMANENT) -> Tuple[List[str], List[str]]:
    full_r1, full_r2 = load_att_matrix_headers_from_template(template_path, "att_matrix")
    perm_r1, perm_r2 = full_r1[:first15], full_r2[:first15]
    dyn_r1_all, dyn_r2_all = full_r1[first15:], full_r2[first15:]
    lodm_norms = lodm_all_attr_norms(lodm_df)
    dyn_r1_kept, dyn_r2_kept = [], []
    for i, code in enumerate(dyn_r1_all):
        if _norm_att_code(code) in lodm_norms:
            dyn_r1_kept.append(code)
            dyn_r2_kept.append(dyn_r2_all[i])
    return perm_r1 + dyn_r1_kept, perm_r2 + dyn_r2_kept

# ---------------- SmartForms (join2) lookup ----------------
def _column_name_candidates_for_code(code: str) -> List[str]:
    base = str(code or "")
    cand = {base, base.replace(":", "_"), base.replace("_", ":"), _RE_SEP.sub("", base)}
    cand |= {x.lower() for x in list(cand)}
    cand |= {_RE_SEP.sub("", x) for x in list(cand)}
    return list(cand)

def ensure_join2_wide(join2_df: pd.DataFrame, logger: Optional[logging.Logger] = None) -> Tuple[pd.DataFrame, Optional[str]]:
    if join2_df is None or join2_df.empty:
        return pd.DataFrame(), None
    uaid3_col = next((c for c in join2_df.columns if _norm_key(c) in {"uaid3","uaid","assetid"}), None)
    if uaid3_col and len(join2_df.columns) > 50:
        return join2_df.astype(str).fillna(""), uaid3_col
    cols = [str(c).strip() for c in join2_df.columns]
    attname_col = next((c for c in cols if normalize_text(c) in {"atttypename","attrtypename","attrtypedisplayname"}), None)
    attid_col   = next((c for c in cols if normalize_text(c) in {"attributetypeid","attrtypeid","atttypeid"}), None)
    val_col     = next((c for c in cols if normalize_text(c) in {"attributevalue","attrvalue","value"}), None)
    if uaid3_col and val_col and (attname_col or attid_col):
        df = join2_df[[uaid3_col, attname_col or attid_col, val_col]].copy()
        df.columns = [uaid3_col, "attr_key", "attr_val"]
        piv = df.pivot_table(index=uaid3_col, columns="attr_key", values="attr_val", aggfunc="first").reset_index()
        piv = piv.astype(str).replace("nan","")
        if logger:
            logger.info("SmartForms L3 pivoted: rows=%d cols=%d", len(piv), len(piv.columns))
        return piv, uaid3_col
    return join2_df.astype(str).fillna(""), uaid3_col

def _pick_join2_uaid3_col(join2_wide: pd.DataFrame) -> str | None:
    if join2_wide is None or join2_wide.empty:
        return None
    targets = {_norm_key(x) for x in ["UAID_3", "UAID3", "UAID", "Uaid", "Asset_ID", "AssetID"]}
    return _first_col_by_keys(join2_wide, targets)

def lookup_join2_value(join2_wide: pd.DataFrame, key_col: str, key_val: str, au_code: str, code_to_id: Dict[str, str]) -> str:
    if join2_wide is None or join2_wide.empty or not key_col or not key_val:
        return ""
    row = join2_wide[join2_wide[key_col].fillna("").astype(str).str.strip() == str(key_val).strip()]
    if row.empty:
        return ""
    norm_map: Dict[str, str] = {}
    for c in row.columns:
        norm_map[_norm_att_code(c)] = c
        norm_map[normalize_text(c)] = c
    for cand in _column_name_candidates_for_code(au_code):
        k1 = _norm_att_code(cand)
        k2 = normalize_text(cand)
        if k1 in norm_map:
            v = row.iloc[0][norm_map[k1]]
            return "" if pd.isna(v) else str(v).strip()
        if k2 in norm_map:
            v = row.iloc[0][norm_map[k2]]
            return "" if pd.isna(v) else str(v).strip()
    att_id = code_to_id.get(_norm_att_code(au_code))
    for alt in (att_id, str(att_id) if att_id else None):
        if alt and alt in row.columns:
            v = row.iloc[0][alt]
            return "" if pd.isna(v) else str(v).strip()
    if att_id and str(att_id).isdigit():
        int_id = int(str(att_id))
        if int_id in row.columns:
            v = row.iloc[0][int_id]
            return "" if pd.isna(v) else str(v).strip()
    return ""

# ---------------- Scope3 UAID columns ----------------
def _pick_scope3_cols(scope3: pd.DataFrame) -> tuple[str | None, str | None]:
    if scope3 is None or scope3.empty:
        return None, None
    u2_keys = {_norm_key(x) for x in ["UAID_2","UAID2","Level 2 UAID","Level2UAID","ParentUaid","Parent_UAID","Parent UAID"]}
    u3_keys = {_norm_key(x) for x in ["UAID_3","UAID3","UAID","Asset_ID","AssetID","Level 3 UAID","Level3UAID"]}
    uaid2_col = _first_col_by_keys(scope3, u2_keys)
    uaid3_col = _first_col_by_keys(scope3, u3_keys)
    return uaid2_col, uaid3_col

# ---------------- Formatting helpers ----------------
def _copy_row_format(ws, source_row: int, target_row: int, max_col: int) -> None:
    if target_row == source_row:
        return
    ws.row_dimensions[target_row].height = ws.row_dimensions[source_row].height
    for col_idx in range(1, max_col + 1):
        src = ws.cell(row=source_row, column=col_idx)
        dst = ws.cell(row=target_row, column=col_idx)
        if src.has_style:
            dst._style = copy.copy(src._style)
        if src.number_format:
            dst.number_format = src.number_format
        if src.alignment:
            dst.alignment = copy.copy(src.alignment)
        if src.protection:
            dst.protection = copy.copy(src.protection)
        if src.hyperlink:
            dst._hyperlink = copy.copy(src.hyperlink)
        if src.comment:
            dst.comment = copy.copy(src.comment)

def _snapshot_row_fills(ws, template_row: int, max_col: int) -> Dict[int, Any]:
    return {
        col_idx: copy.copy(ws.cell(row=template_row, column=col_idx).fill)
        for col_idx in range(1, max_col + 1)
    }

def _blank_fill_from_snapshot(ws, template_fills: Dict[int, Any], target_row: int, col_idx: int) -> None:
    fill = template_fills.get(col_idx)
    if fill is not None:
        ws.cell(row=target_row, column=col_idx).fill = copy.copy(fill)

# ---------------- Applicability ----------------
def _build_row_allowed_columns(
    mpdt_df: pd.DataFrame,
    columns: List[str],
    lodm_df: pd.DataFrame,
    aw_start_idx: int,
    logger: Optional[logging.Logger] = None,
    row1_codes: Optional[Dict[str, str]] = None,
) -> Tuple[List[Set[str]], Set[str]]:
    class_col = next((c for c in columns if _norm_key(c) in {"assethierarchycategory","hs2class","classcode"}), None)
    aw_columns = [c for i, c in enumerate(columns, start=1) if i >= aw_start_idx]
    row_sets: List[Set[str]] = []
    union_set: Set[str] = set()
    if mpdt_df is None or mpdt_df.empty or lodm_df is None or lodm_df.empty:
        return [set() for _ in range(0 if mpdt_df is None else len(mpdt_df))], union_set
    header_to_code_norm: Dict[str, str] = {}
    if row1_codes:
        for h2, h1 in row1_codes.items():
            header_to_code_norm[str(h2).strip()] = _norm_att_code(h1)
    for _, row in mpdt_df.iterrows():
        cc = str(row.get(class_col, "")).strip() if class_col else ""
        allowed_codes = lodm_allowed_codes_for_class(lodm_df, cc)
        allowed_cols: Set[str] = set()
        for h in aw_columns:
            code_norm = header_to_code_norm.get(h, "")
            if code_norm and code_norm in allowed_codes:
                allowed_cols.add(h)
        row_sets.append(allowed_cols)
        union_set |= allowed_cols
    if logger and logger.isEnabledFor(logging.INFO):
        logger.info("    Applicability | %d rows | union allowed cols: %d / %d", len(row_sets), len(union_set), len(aw_columns))
    return row_sets, union_set

def apply_applicability(
    ws,
    data_rows: Sequence[Mapping[str, Any]],
    columns: List[str],
    row_allowed_cols: List[Set[str]],
    start_row: int = 3,
    union_allowed_cols: Optional[Set[str]] = None,
    template_fills: Optional[Dict[int, Any]] = None,
) -> None:
    aw_start_idx = column_index_from_string("AW") if column_index_from_string else 49
    if aw_start_idx > len(columns):
        return
    for row_offset, _row_data in enumerate(data_rows):
        excel_row = start_row + row_offset
        allowed = row_allowed_cols[row_offset] if row_offset < len(row_allowed_cols) else set()
        for col_idx, col in enumerate(columns, start=1):
            if col_idx < aw_start_idx:
                if template_fills is not None:
                    _blank_fill_from_snapshot(ws, template_fills, excel_row, col_idx)
                continue
            is_class_attr = union_allowed_cols is None or col in union_allowed_cols
            if is_class_attr and col not in allowed:
                if _BLACK_FILL is not None:
                    ws.cell(row=excel_row, column=col_idx).fill = _BLACK_FILL
            else:
                if template_fills is not None:
                    _blank_fill_from_snapshot(ws, template_fills, excel_row, col_idx)

# ---------------- Cached mapping expr + RowResolver ----------------
_JOIN_EXPR_RE = re.compile(r"^(?:join|join_)?([12])\s*$$\s*['\"]([^'\"]+)['\"]\s*$$$", re.IGNORECASE)

@lru_cache(maxsize=None)
def _parse_mapping_expr_cached(expr: str) -> Tuple[str, Optional[int], str]:
    s = str(expr or "").strip()
    m = _JOIN_EXPR_RE.match(s)
    if m:
        return ("join", int(m.group(1)), m.group(2))
    return ("literal", None, s)

@dataclass
class RowResolver:
    exact: Dict[str, Any]
    nosuf: Dict[str, Any]

    @classmethod
    def from_series(cls, row: Optional[pd.Series], suffixes: Tuple[str, ...] = ("_sform2", "_sform3", "_l2", "_l3")):
        if row is None or row.empty:
            return cls({}, {})
        exact: Dict[str, Any] = {}
        nosuf: Dict[str, Any] = {}
        for col, val in row.items():
            if _is_empty(val):
                continue
            raw = str(col)
            nraw = normalize_text(raw)
            exact.setdefault(nraw, val)
            base = raw
            lower = raw.lower()
            for suf in suffixes:
                if lower.endswith(suf):
                    base = raw[: -len(suf)]
                    break
            nosuf.setdefault(normalize_text(base), val)
        return cls(exact, nosuf)

    def get(self, colname: str) -> Any:
        key = normalize_text(colname)
        return self.exact.get(key) if key in self.exact else self.nosuf.get(key)

def _eval_mapping_expr(
    expr: Any,
    join2_resolver: Optional[RowResolver],
    join1_resolver: Optional[RowResolver],
    deliverable_name: str = "",
) -> Any:
    if expr is None:
        return ""
    kind, join_idx, col = _parse_mapping_expr_cached(str(expr))
    if kind == "join" and col:
        if join_idx == 2 and join2_resolver:
            v = join2_resolver.get(col)
            return "" if _is_empty(v) else v
        if join_idx == 1 and join1_resolver:
            v = join1_resolver.get(col)
            return "" if _is_empty(v) else v
        return ""
    if kind == "literal":
        lit = col
        if lit.lower() in {"deliverablename","deliverable_name","deliverable","filename"}:
            return deliverable_name
        if join2_resolver:
            v = join2_resolver.get(lit)
            if not _is_empty(v):
                return v
        if join1_resolver:
            v = join1_resolver.get(lit)
            if not _is_empty(v):
                return v
        return lit if lit else ""
    return ""

# ---------------- AU debug ----------------
def _debug_probe_au_row(logger: logging.Logger, uaid3_col_j2: str, uaid3_val: str, au_row1: List[str], join2_wide: pd.DataFrame, code_to_id: Dict[str, str]) -> None:
    if not logger.isEnabledFor(logging.DEBUG):
        return
    logger.debug("  AU probe | UAID_3 col=%s value=%s | AU codes=%d", uaid3_col_j2, uaid3_val, len(au_row1))
    row = pd.DataFrame()
    if join2_wide is not None and not join2_wide.empty and uaid3_col_j2:
        row = join2_wide[join2_wide[uaid3_col_j2].fillna("").astype(str).str.strip() == str(uaid3_val).strip()]
    if row.empty:
        logger.debug("  AU probe | No SmartForms row matched UAID_3 — AU+ will be blank")
        return
    found, missing = [], []
    for code in au_row1[:24]:
        v = lookup_join2_value(join2_wide, uaid3_col_j2, uaid3_val, code, code_to_id)
        (found if v else missing).append(code)
    logger.debug("  AU probe | Found values for: %s", ", ".join(found) if found else "(none)")
    logger.debug("  AU probe | No values for: %s", ", ".join(missing) if missing else "(none)")

# ---------------- Core per-UAID generator ----------------
def generate_single_mpdt(
    workspace: Path,
    cfg: dict,
    uaid2: str,
    mpdt_dir: Path,
    mapping_dict: dict,
    columns: list[str],
    scope2: pd.DataFrame,
    scope3: pd.DataFrame,
    sf_l2: pd.DataFrame,
    sf_l3: pd.DataFrame,
    lodm_df: pd.DataFrame,
    pw_df: pd.DataFrame,
    control_df: pd.DataFrame,
    logger: logging.Logger,
    deliverable_file: str,
    l2_df: pd.DataFrame | None = None,
    row1_codes: dict | None = None,
    indexes: dict | None = None,
) -> Path | None:
    logger.info("  Generating MPDT for %s", uaid2)

    tpl_rel = cfg.get("paths", {}).get("mpdt_template", "Input/C2-MPDT-Template-Mapping.xlsm")
    template_path = (workspace / tpl_rel).resolve()
    if not template_path.exists():
        logger.error("  MPDT template not found: %s", template_path)
        return None

    if deliverable_file:
        stem = Path(deliverable_file).stem
        filename = f"{stem}.xlsm"
    else:
        filename = f"MPDT_{uaid2}.xlsm"
    mpdt_dir.mkdir(parents=True, exist_ok=True)
    output_file = get_available_path(mpdt_dir / sanitize_filename(filename))

    # Copy template and load
    import shutil
    shutil.copy(template_path, output_file)
    wb = load_workbook(str(output_file), keep_vba=True)
    ws = wb["MPDT Element of Asset"] if "MPDT Element of Asset" in wb.sheetnames else wb[wb.sheetnames[0]]

    if not isinstance(columns, list) or len(columns) == 0:
        logger.error("  'columns' not provided.")
        return None

    # Filter scope3 by UAID_2
    uaid2_col_s3, uaid3_col_s3 = _pick_scope3_cols(scope3)
    if not uaid2_col_s3:
        logger.error("  UAID_2 column not found in scope3; available: %s", list(scope3.columns)[:20])
        wb.save(output_file)
        return None
    rows_for_uaid = scope3[scope3[uaid2_col_s3].fillna("").astype(str).str.strip().str.upper() == str(uaid2).strip().upper()]
    if rows_for_uaid.empty:
        logger.warning("  No scope3 rows for UAID_2=%s — skipping", uaid2)
        wb.save(output_file)
        return None

    # SmartForms L3 wide + UAID_3 key
    join2_wide, uaid3_col_j2 = ensure_join2_wide(sf_l3, logger=logger)
    if not uaid3_col_j2 and not join2_wide.empty:
        uaid3_col_j2 = _pick_join2_uaid3_col(join2_wide)
    logger.info("  Keys | scope3: UAID_2=%s UAID_3=%s | join2: UAID_3=%s", uaid2_col_s3, uaid3_col_s3, uaid3_col_j2)

    # join1 row (scope2) for UAID_2
    join1_row = None
    if scope2 is not None and not scope2.empty:
        u2_col_s2 = next((c for c in scope2.columns if _norm_key(c) in {"uaid2","uaid"}), None)
        if u2_col_s2:
            sub1 = scope2[scope2[u2_col_s2].fillna("").astype(str).str.strip().str.upper() == str(uaid2).strip().upper()]
            if not sub1.empty:
                join1_row = sub1.iloc[0]

    # Build data_rows using resolvers
    data_rows: List[Dict[str, Any]] = []
    deliverable_name = Path(deliverable_file).stem if deliverable_file else f"MPDT_{uaid2}"
    for _, s3 in rows_for_uaid.iterrows():
        # join2 row for UAID_3
        join2_row = None
        u3_val = str(s3.get(uaid3_col_s3 or "UAID_3", "")).strip() if (uaid3_col_s3 or "UAID_3" in s3.index) else ""
        if join2_wide is not None and not join2_wide.empty and uaid3_col_j2 and u3_val:
            sub2 = join2_wide[join2_wide[uaid3_col_j2].fillna("").astype(str).str.strip() == u3_val]
            if not sub2.empty:
                join2_row = sub2.iloc[0]

        join2_res = RowResolver.from_series(join2_row)
        join1_res = RowResolver.from_series(join1_row)

        row_dict: Dict[str, Any] = {}
        for col in columns:
            expr = mapping_dict.get(col) or mapping_dict.get(col.strip())
            val = _eval_mapping_expr(expr, join2_res, join1_res, deliverable_name)
            if _is_empty(val) and col in s3.index:
                val = s3.get(col, "")
            row_dict[str(col)] = val
        data_rows.append(row_dict)

    # Local DF for downstream
    mpdt_local_df = pd.DataFrame(data_rows, columns=columns)
    logger.info("  %s: constructed %d data rows from scope3", uaid2, len(mpdt_local_df))

    # Clear sheet rows to be written
    last_needed_row = DATA_FIRST_ROW + max(0, len(data_rows) - 1)
    for r in range(DATA_FIRST_ROW, last_needed_row + 1):
        for c in range(1, ws.max_column + 1):
            ws.cell(row=r, column=c).value = None

    # Write governance rows
    template_data_row = DATA_FIRST_ROW
    gov_records: List[Dict[str, Any]] = [{str(k): v for k, v in rec.items()} for rec in mpdt_local_df.to_dict("records")]
    for row_offset, row_data in enumerate(gov_records):
        excel_row = DATA_FIRST_ROW + row_offset
        _copy_row_format(ws, template_data_row, excel_row, ws.max_column)
        for col_idx, col in enumerate(columns, start=1):
            ws.cell(row=excel_row, column=col_idx).value = row_data.get(col)

    # AU+ headers from att_matrix
    au_start_idx = column_index_from_string(DEFAULT_AU_LETTER) if column_index_from_string else _col_letter_to_index(DEFAULT_AU_LETTER)
    au_row1, au_row2 = build_final_au_headers(template_path, lodm_df, FIRST15_PERMANENT)
    for off, (h1, h2) in enumerate(zip(au_row1, au_row2)):
        ws.cell(row=HEADER_ROW1, column=au_start_idx + off, value=h1)
        ws.cell(row=HEADER_ROW2, column=au_start_idx + off, value=h2)

    # LoDM maps
    code_to_id, _class_to_codes = build_lodm_maps(lodm_df)

    # Clear AU region
    total_rows = len(mpdt_local_df)
    if total_rows > 0:
        for rr in range(DATA_FIRST_ROW, DATA_FIRST_ROW + total_rows):
            for cc in range(au_start_idx, au_start_idx + len(au_row1)):
                ws.cell(row=rr, column=cc, value=None)

    # Probe first row AU join
    if total_rows > 0 and uaid3_col_j2 and uaid3_col_s3:
        probe_u3 = str(rows_for_uaid.iloc[0].get(uaid3_col_s3, "")).strip()
        if probe_u3:
            _debug_probe_au_row(logger, uaid3_col_j2, probe_u3, au_row1, join2_wide, code_to_id)

    # Fill AU+ and Count
    count_col_idx = None
    for c in range(1, ws.max_column + 1):
        val = ws.cell(row=HEADER_ROW2, column=c).value
        if val is not None and normalize_text(val) == "count":
            count_col_idx = c
            break

    for row_offset in range(total_rows):
        excel_row = DATA_FIRST_ROW + row_offset
        u3_val = str(rows_for_uaid.iloc[row_offset].get(uaid3_col_s3, "")).strip() if uaid3_col_s3 else ""
        for off, code in enumerate(au_row1):
            v = lookup_join2_value(join2_wide, uaid3_col_j2, u3_val, code, code_to_id) if uaid3_col_j2 else ""
            ws.cell(row=excel_row, column=au_start_idx + off, value=v)
        class_code = str(
            mpdt_local_df.iloc[row_offset].get("HS2_Class")
            or mpdt_local_df.iloc[row_offset].get("AssetHierarchyCategory")
            or rows_for_uaid.iloc[row_offset].get("HS2_Class")
            or rows_for_uaid.iloc[row_offset].get("AssetHierarchyCategory")
            or ""
        ).strip()
        if count_col_idx:
            ws.cell(row=excel_row, column=count_col_idx, value=lodm_attr_count_for_class(lodm_df, class_code) + 10)

    # Applicability
    aw_start_idx = column_index_from_string("AW") if column_index_from_string else 49
    if isinstance(row1_codes, dict):
        for h1, h2 in zip(au_row1, au_row2):
            if str(h2).strip():
                row1_codes[str(h2).strip()] = str(h1).strip()
    try:
        row_allowed_cols, union_allowed_cols = _build_row_allowed_columns(
            mpdt_local_df,
            columns,
            lodm_df,
            aw_start_idx,
            logger,
            row1_codes=row1_codes,
        )
        template_fills = _snapshot_row_fills(ws, DATA_FIRST_ROW, ws.max_column)
        apply_applicability(
            ws,
            gov_records,  # Sequence[Mapping[str, Any]]
            columns,
            row_allowed_cols,
            start_row=DATA_FIRST_ROW,
            union_allowed_cols=union_allowed_cols,
            template_fills=template_fills,
        )
    except Exception as ex:
        logger.warning("  Applicability pass skipped: %s", ex)

    # Save
    wb.save(output_file)
    logger.info("  MPDT written: %s (%d data rows)", output_file.name, len(mpdt_local_df))
    return output_file

# ---------------- Batch wrapper with precomputes ----------------
def _pick_df(sources: dict, primary_key: str, fallback_key: str) -> pd.DataFrame:
    df = sources.get(primary_key, None)
    if df is None or (hasattr(df, "empty") and df.empty):
        fb = sources.get(fallback_key, None)
        return fb if fb is not None else pd.DataFrame()
    return df

def generate_mpdt_batch(
    workspace: Path,
    cfg: dict,
    targets: list[dict],
    output_dir: Path,
    sources: dict,
    logger: logging.Logger,
) -> dict:
    scope3 = sources.get("scope3_df", pd.DataFrame())
    if scope3 is None or scope3.empty:
        raise RuntimeError("AssetsScope3 not found")

    sf_l3 = _pick_df(sources, "sf_l3_df", "sf_l3")
    scope2 = sources.get("scope2_df", pd.DataFrame())
    sf_l2 = _pick_df(sources, "sf_l2_df", "sf_l2")
    lodm_df = sources.get("lodm_df", pd.DataFrame())
    pw_df = sources.get("pw_df", pd.DataFrame())
    control_df = sources.get("control_df", pd.DataFrame())
    l2_df = sources.get("l2_df", pd.DataFrame())
    mapping_dict = sources.get("mapping_dict", {})
    columns = sources.get("mpdt_columns", [])

    # Batch precomputes (indexes)
    lodm_attr_count_map = {}
    lodm_class_name_map = {}
    if lodm_df is not None and not lodm_df.empty:
        cc_col = next((c for c in lodm_df.columns if normalize_text(c) == "classcode"), None)
        an_col = next((c for c in lodm_df.columns if normalize_text(c) == "atttypename"), None)
        cn_col = next((c for c in lodm_df.columns if normalize_text(c) in {"classname","class name"}), None)
        if cc_col and an_col:
            tmp = lodm_df[[cc_col, an_col]].dropna()
            tmp["__cc__"] = tmp[cc_col].astype(str).str.strip().str.lower()
            tmp["__an__"] = tmp[an_col].astype(str).str.strip()
            grp = tmp.groupby("__cc__")["__an__"].agg(lambda s: len({ _norm_att_code(v) for v in s if str(v).strip()}))
            lodm_attr_count_map = grp.to_dict()
        if cc_col and cn_col:
            tmp2 = lodm_df[[cc_col, cn_col]].drop_duplicates(subset=[cc_col])
            lodm_class_name_map = { str(k).strip().lower(): v for k, v in zip(tmp2[cc_col], tmp2[cn_col]) }

    control_desc_map = {}
    control_disc_map = {}
    if control_df is not None and not control_df.empty:
        key_col = next((c for c in control_df.columns if normalize_text(c) in {"classcode","hs2_class"}), None)
        desc_col = next((c for c in control_df.columns if "description" in normalize_text(c)), None)
        disc_col = next((c for c in control_df.columns if "discipline" in normalize_text(c)), None)
        if key_col and desc_col:
            k = control_df[key_col].astype(str).str.strip().str.lower()
            control_desc_map = { kk: vv for kk, vv in zip(k, control_df[desc_col]) }
        if key_col and disc_col:
            k = control_df[key_col].astype(str).str.strip().str.lower()
            control_disc_map = { kk: vv for kk, vv in zip(k, control_df[disc_col]) }

    latest_pw_row_by_uaid: Dict[str, pd.Series] = {}
    if pw_df is not None and not pw_df.empty:
        uaid_candidates = [c for c in pw_df.columns if normalize_text(c) in {"pw_uaid","asset_id","uaid_2","zz_lbl_uaidl2","zz_lbl_uaidl1","uaid"}]
        date_col = None
        for c in ("FileUpdated_parsed","FileUpdated"):
            if c in pw_df.columns:
                date_col = c
                break
        dfpw = pw_df.copy()
        dfpw["__file_date__"] = pd.to_datetime(dfpw[date_col], errors="coerce") if date_col else pd.NaT
        for c in uaid_candidates:
            key_series = dfpw[c].astype(str).str.strip().str.upper()
            dfpw["__uaid_key__"] = key_series
            sortd = dfpw.sort_values(["__uaid_key__","__file_date__"], ascending=[True, False], na_position="last")
            latest_pw_row_by_uaid.update(sortd.drop_duplicates(subset="__uaid_key__", keep="first").set_index("__uaid_key__").to_dict(orient="index"))

    uaid2_col_s3 = next((c for c in scope3.columns if _norm_key(c) in {"uaid2","parentuaid"}), None)
    uaid3_col_s3 = next((c for c in scope3.columns if _norm_key(c) in {"uaid3","uaid","assetid"}), None)
    if uaid2_col_s3:
        _build_normalized_key(scope3, uaid2_col_s3, "__uaid2_key__")
    if uaid3_col_s3:
        _build_normalized_key(scope3, uaid3_col_s3, "__uaid3_key__")
    sf_l3_u3 = next((c for c in sf_l3.columns if _norm_key(c) in {"uaid3","uaid","assetid"}), None) if sf_l3 is not None and not sf_l3.empty else None
    if sf_l3_u3:
        _build_normalized_key(sf_l3, sf_l3_u3, "__uaid3_key__")

    idx = dict(sources.get("indexes") or {})
    idx.update({
        "lodm_attr_count_map": lodm_attr_count_map,
        "lodm_class_name_map": lodm_class_name_map,
        "control_desc_map": control_desc_map,
        "control_disc_map": control_disc_map,
        "latest_pw_row_by_uaid": latest_pw_row_by_uaid,
        "scope3_uaid2_col": uaid2_col_s3,
        "scope3_uaid3_col": uaid3_col_s3,
        "sf_l3_uaid3_col": sf_l3_u3,
    })
    sources["indexes"] = idx

    mpdt_dir = output_dir / "MPDT"
    mpdt_dir.mkdir(parents=True, exist_ok=True)

    generated, errors = [], []
    for target in targets:
        uaid2 = target["uaid"]
        deliverable_file = target.get("file", "")
        try:
            out = generate_single_mpdt(
                workspace, cfg, uaid2, mpdt_dir,
                mapping_dict, columns,
                scope2, scope3, sf_l2, sf_l3,
                lodm_df, pw_df, control_df,
                logger, deliverable_file,
                l2_df=l2_df,
                row1_codes=sources.get("row1_codes"),
                indexes=sources.get("indexes"),
            )
            if out:
                generated.append({"uaid": uaid2, "file": str(out)})
            else:
                errors.append({"uaid": uaid2, "error": "Generation returned None"})
        except Exception as exc:
            logger.error("MPDT generation failed for %s: %s", uaid2, exc, exc_info=True)
            errors.append({"uaid": uaid2, "error": str(exc)})

    return {"generated": generated, "errors": errors}

# ---------------- Standalone (optional, for local debug) ----------------
def main() -> None:
    parser = argparse.ArgumentParser(description="Generate MPDT files (att_matrix AU+ patch)")
    parser.add_argument("--workspace", default=None)
    parser.add_argument("--target-uaid2", nargs="+", required=True)
    args = parser.parse_args()

    workspace = resolve_workspace(args.workspace)
    cfg = load_config(workspace)
    logger = setup_logger(workspace, "mpdt_generator", cfg.get("log_level", "INFO"))
    logger.info("=== MPDT Generator (att_matrix AU+) ===")

    paths = cfg.get("paths", {})
    sources: Dict[str, pd.DataFrame] = {}
    try: sources["scope3_df"] = read_table_any(workspace / paths.get("l3_assets_scope_data", "Input/l3_assets_scope_data.xlsx"))
    except Exception: sources["scope3_df"] = pd.DataFrame()
    try: sources["sf_l3_df"] = read_table_any(workspace / paths.get("smartforms_fallback", "Input/SmartForms_RAW_MPDT_L2&L3.xlsx"))
    except Exception: sources["sf_l3_df"] = pd.DataFrame()
    try: sources["lodm_df"] = read_table_any(workspace / paths.get("lodm", "Input/1MC06-ASC-IM-SCH-C002-000009_lodm.xlsx"))
    except Exception: sources["lodm_df"] = pd.DataFrame()
    try: sources["pw_df"] = read_table_any(workspace / paths.get("pw_extract", "Input/ACBOS_MPDT.xlsx"))
    except Exception: sources["pw_df"] = pd.DataFrame()
    try: sources["control_df"] = read_table_any(workspace / paths.get("control_file", "Input/1MC06-ASC-IM-GDE-C002-000090_controlfile.xlsx"))
    except Exception: sources["control_df"] = pd.DataFrame()
    try: sources["l2_df"] = read_table_any(workspace / paths.get("l2_uaid_acbos", "Input/L2 UAID-ACBOS.xlsx"))
    except Exception: sources["l2_df"] = pd.DataFrame()

    uaids = []
    for v in args.target_uaid2:
        uaids.extend(u.strip() for u in v.split(",") if u.strip())
    targets = [{"uaid": u} for u in uaids]

    output_dir = timestamped_dir(workspace, "Output")
    result = generate_mpdt_batch(workspace, cfg, targets, output_dir, sources, logger)
    write_json(workspace / "Output" / "mpdt_result.json", result)
    logger.info("Generated: %d, Errors: %d", len(result["generated"]), len(result["errors"]))

if __name__ == "__main__":
    main()