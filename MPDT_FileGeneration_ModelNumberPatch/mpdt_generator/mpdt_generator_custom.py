#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
mpdt_generator_custom.py — Generate MPDT Excel files per UAID_2 using att_matrix for AU+ columns
and LoDM-driven filtering, with checkerboard applicability and LoDM-based count.

Implements the following:
- AU row1 from template att_matrix column C (AttType short code)
- AU row2 from template att_matrix "E B" (display label assembly)
- First 15 AU columns permanent (always kept)
- Remaining AU columns filtered to those present in LoDM (any class)
- Populate AU cells from join2_df (wide) by AttType code for each row
- Checkerboard blackening for non-applicable attributes (based on row's ClassCode in LoDM)
- Count = LoDM attribute count for row's class + 10

References:
- LoDM attribute mapping and normalization approach [2.1, 2.2, 2.3, 2.4]
- SmartForms pivot to wide for join2_df [5.3]
"""

from __future__ import annotations

import argparse
import logging
import re
from pathlib import Path
from typing import Dict, Iterable, List, Optional, Sequence, Set, Tuple

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# ----------------------------
# Configurable constants
# ----------------------------

FIRST15_PERMANENT = 15
HEAD_ROW1 = 1   # Row index holding AU Row 1 header (att short code)
HEAD_ROW2 = 2   # Row index holding AU Row 2 header (display label)
DATA_FIRST_ROW = 3  # First row for data
DEFAULT_AU_LETTER = "AU"  # Used if AU detection fails
DEFAULT_COUNT_HEADER = "count"  # Row-2 label for the count column (case-insensitive)


# ----------------------------
# Utilities and normalization
# ----------------------------

_RE_SEP = re.compile(r'[\s_\-\.:]+')

def col_letter_to_index(letter: str) -> int:
    """Convert Excel column letter to 1-based index."""
    letter = letter.strip().upper()
    num = 0
    for ch in letter:
        if not ('A' <= ch <= 'Z'):
            continue
        num = num * 26 + (ord(ch) - ord('A') + 1)
    return num

def normalize_text(s: str) -> str:
    return re.sub(r"\s+", " ", str(s or "").strip()).lower()

def norm_att_code(s: str) -> str:
    """Normalize an AttTypeName/short code by removing separators and lowercasing, e.g. 'Com_Dscrptn' = 'Com:Dscrptn'."""
    return _RE_SEP.sub('', str(s or "")).lower()

def norm_code_variants(code: str) -> Set[str]:
    """LoDM ClassCode variants to improve matching robustness."""
    s = str(code or "").strip()
    if not s:
        return set()
    c = s.lower()
    return {c, c.replace("_", "-"), c.replace("-", "_")}

def is_empty(value) -> bool:
    if value is None:
        return True
    try:
        if pd.isna(value):
            return True
    except Exception:
        pass
    return str(value).strip() in ("", "nan", "none", "nat", "NaT", "NaN")


# ----------------------------
# Loaders
# ----------------------------

def read_table_any(path: Path) -> pd.DataFrame:
    if not path or not Path(path).exists():
        return pd.DataFrame()
    ext = Path(path).suffix.lower()
    if ext == ".csv":
        return pd.read_csv(path, dtype=str)
    if ext in (".xlsx", ".xlsm", ".xls"):
        return pd.read_excel(path, dtype=str)
    return pd.DataFrame()

def ensure_join2_wide(join2_df: pd.DataFrame, uaid3_col_candidates: Sequence[str] = ("UAID_3", "Uaid_3", "Uaid", "UAID", "Asset_ID")) -> Tuple[pd.DataFrame, Optional[str]]:
    """
    Ensure join2_df is in wide form with a primary key column (UAID_3).
    If it's already wide, just return it. If it's normalized (attr rows), try pivoting (see [5.3]).
    """
    if join2_df is None or join2_df.empty:
        return pd.DataFrame(), None

    # Detect UAID_3 column
    uaid3_col = None
    for c in join2_df.columns:
        if normalize_text(c) in {normalize_text(k) for k in uaid3_col_candidates}:
            uaid3_col = c
            break

    # If join2_df already has many attribute columns, assume it's wide
    # Heuristic: if > 50 columns and has UAID_3, treat as wide
    if uaid3_col and len(join2_df.columns) > 50:
        return join2_df.astype(str).fillna(""), uaid3_col

    # Try pivot based on typical normalized columns
    cols = [str(c).strip() for c in join2_df.columns]
    attr_id_col = next((c for c in cols if normalize_text(c) in {"atttypename", "attrtypename", "attrtypeid", "attrtypedisplayname"}), None)
    val_col = next((c for c in cols if normalize_text(c) in {"attributevalue", "attrvalue", "value"}), None)

    if uaid3_col and attr_id_col and val_col:
        wide = join2_df[[uaid3_col, attr_id_col, val_col]].copy()
        wide.columns = [uaid3_col, "attr_id", "attr_val"]
        pivoted = wide.pivot_table(index=uaid3_col, columns="attr_id", values="attr_val", aggfunc="first")
        pivoted = pivoted.reset_index()
        pivoted = pivoted.astype(str).replace("nan", "")
        return pivoted, uaid3_col

    # Fallback: best-effort
    return join2_df.astype(str).fillna(""), uaid3_col


# ----------------------------
# att_matrix -> AU headers
# ----------------------------

def load_att_matrix_headers_from_template(template_path: Path, sheet_name: str = "att_matrix") -> Tuple[List[str], List[str]]:
    """
    Returns (au_row1_list, au_row2_list):
    - Row 1 values from att_matrix column C
    - Row 2 values from "E B" (E then B)
    """
    wb = load_workbook(template_path, read_only=True, data_only=True, keep_vba=True)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"att_matrix sheet not found in template: {sheet_name}")
    ws = wb[sheet_name]
    r1: List[str] = []
    r2: List[str] = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        # Columns: A=0, B=1, C=2, D=3, E=4
        col_b = str(r[1] or "").strip()
        col_c = str(r[2] or "").strip()
        col_e = str(r[4] or "").strip()
        if not col_c:
            continue
        row1 = col_c
        row2 = " ".join([x for x in (col_e, col_b) if str(x).strip()])
        r1.append(row1)
        r2.append(row2)
    return r1, r2

def build_final_au_headers(template_path: Path, lodm_df: pd.DataFrame, first15: int = FIRST15_PERMANENT) -> Tuple[List[str], List[str]]:
    """
    Build AU Row1 and Row2:
    - First 15 from att_matrix kept permanently
    - Remaining items filtered to those present in LoDM (any class) based on AttType short code normalization
    """
    full_r1, full_r2 = load_att_matrix_headers_from_template(template_path, "att_matrix")
    perm_r1 = full_r1[:first15]
    perm_r2 = full_r2[:first15]
    dyn_r1_all = full_r1[first15:]
    dyn_r2_all = full_r2[first15:]

    lodm_norms = lodm_all_attr_norms(lodm_df)  # [2.2]
    dyn_r1_kept: List[str] = []
    dyn_r2_kept: List[str] = []
    for i, code in enumerate(dyn_r1_all):
        if norm_att_code(code) in lodm_norms:
            dyn_r1_kept.append(code)
            dyn_r2_kept.append(dyn_r2_all[i])

    return perm_r1 + dyn_r1_kept, perm_r2 + dyn_r2_kept


# ----------------------------
# LoDM helpers
# ----------------------------

def lodm_all_attr_norms(lodm_df: pd.DataFrame) -> Set[str]:
    """Normalized set of all LoDM AttTypeName values across classes [2.2]."""
    if lodm_df is None or lodm_df.empty:
        return set()
    att_col = next((c for c in lodm_df.columns if normalize_text(c) == "atttypename"), None)
    if not att_col:
        return set()
    return {norm_att_code(v) for v in lodm_df[att_col].dropna().astype(str) if str(v).strip()}

def lodm_allowed_codes_for_class(lodm_df: pd.DataFrame, class_code: str) -> Set[str]:
    """Normalized AttType short codes allowed for a class [2.6, 2.2]."""
    if lodm_df is None or lodm_df.empty or not class_code:
        return set()
    cc_col = next((c for c in lodm_df.columns if normalize_text(c) == "classcode"), None)
    att_col = next((c for c in lodm_df.columns if normalize_text(c) == "atttypename"), None)
    if not cc_col or not att_col:
        return set()
    variants = norm_code_variants(class_code)
    rows = lodm_df[lodm_df[cc_col].fillna("").astype(str).str.strip().str.lower().isin(variants)]
    return {norm_att_code(v) for v in rows[att_col].dropna().astype(str) if str(v).strip()}

def lodm_attr_count_for_class(lodm_df: pd.DataFrame, class_code: str) -> int:
    """Number of LoDM attributes for a classCode (normalized unique AttTypeName) [2.3]."""
    return len(lodm_allowed_codes_for_class(lodm_df, class_code))


# ----------------------------
# AU header write + value population
# ----------------------------

def write_au_headers(ws, au_col_idx: int, au_row1: List[str], au_row2: List[str]) -> None:
    """Write Row1 and Row2 headers for AU+."""
    for off, (h1, h2) in enumerate(zip(au_row1, au_row2)):
        c = au_col_idx + off
        ws.cell(row=HEAD_ROW1, column=c, value=h1)
        ws.cell(row=HEAD_ROW2, column=c, value=h2)

def _column_name_candidates_for_code(code: str) -> List[str]:
    """Generate candidate column names for a join2 wide DataFrame given a Row1 short code."""
    variants = set()
    base = str(code or "")
    variants.add(base)
    # Replace separators both ways
    variants.add(base.replace(":", "_"))
    variants.add(base.replace("_", ":"))
    # Add fully normalized (no separators)
    variants.add(_RE_SEP.sub("", base))
    # Lowercase variants
    lower_variants = set()
    for v in list(variants):
        lower_variants.add(v.lower())
        lower_variants.add(_RE_SEP.sub("", v.lower()))
    return list(variants | lower_variants)

def lookup_join2_value(join2_wide: pd.DataFrame, key_col: str, key_val: str, code: str) -> str:
    """Look up value in join2_wide at (key_val, code) with tolerant column matching."""
    if join2_wide is None or join2_wide.empty or not key_col:
        return ""
    if is_empty(key_val):
        return ""
    row = join2_wide[join2_wide[key_col].fillna("").astype(str).str.strip() == str(key_val).strip()]
    if row.empty:
        return ""
    # Exact column
    if code in row.columns:
        v = row.iloc[0][code]
        return "" if pd.isna(v) else str(v).strip()
    # Tolerant match
    cand_cols = _column_name_candidates_for_code(code)
    # Build a mapping of normalized name -> real column
    col_norm_map: Dict[str, str] = {}
    for c in row.columns:
        col_norm_map[norm_att_code(c)] = c
        col_norm_map[normalize_text(c)] = c
    for cand in cand_cols:
        # try normalized no-sep
        key = norm_att_code(cand)
        if key in col_norm_map:
            real = col_norm_map[key]
            v = row.iloc[0][real]
            return "" if pd.isna(v) else str(v).strip()
        # try normalized text
        key2 = normalize_text(cand)
        if key2 in col_norm_map:
            real = col_norm_map[key2]
            v = row.iloc[0][real]
            return "" if pd.isna(v) else str(v).strip()
    return ""

def fill_au_values_row(ws, row_idx: int, au_col_idx: int, au_row1: List[str], join2_wide: pd.DataFrame, key_col: str, key_val: str) -> None:
    for off, code in enumerate(au_row1):
        val = lookup_join2_value(join2_wide, key_col, key_val, code)
        ws.cell(row=row_idx, column=au_col_idx + off, value=val)


# ----------------------------
# Checkerboard and Count
# ----------------------------

def apply_checkerboard(ws, row_idx: int, au_col_idx: int, au_row1: List[str], class_code: str, lodm_df: pd.DataFrame, first15: int = FIRST15_PERMANENT) -> None:
    """Black fill AU cells beyond first15 not applicable for the row's class [2.2, 2.4]."""
    if lodm_df is None or lodm_df.empty:
        return
    allowed = lodm_allowed_codes_for_class(lodm_df, class_code)
    black = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
    for off, code in enumerate(au_row1):
        if off < first15:
            continue  # permanent, never blacken
        norm = norm_att_code(code)
        if norm and norm not in allowed:
            ws.cell(row=row_idx, column=au_col_idx + off).fill = black

def write_count(ws, row_idx: int, count_col_idx: int, lodm_df: pd.DataFrame, class_code: str) -> None:
    cnt = lodm_attr_count_for_class(lodm_df, class_code)
    ws.cell(row=row_idx, column=count_col_idx, value=cnt + 10)


# ----------------------------
# Detect AU start and count column
# ----------------------------

def find_au_start(ws) -> Optional[int]:
    """Attempt to detect AU start as first empty cell in row1 after governance block; fallback to DEFAULT_AU_LETTER."""
    # If AU is known by letter, prefer that
    if DEFAULT_AU_LETTER:
        try:
            return col_letter_to_index(DEFAULT_AU_LETTER)
        except Exception:
            pass
    # Fallback naive scan: find first blank cell in row1 after a long non-empty header region
    maxc = ws.max_column
    seen_nonempty = 0
    for c in range(1, maxc + 1):
        v = ws.cell(row=HEAD_ROW1, column=c).value
        if v is not None and str(v).strip():
            seen_nonempty += 1
        elif seen_nonempty >= 5:
            return c
    return None

def find_count_col_idx(ws) -> Optional[int]:
    """Try to find a 'count' column in row2 by label (case-insensitive)."""
    maxc = ws.max_column
    for c in range(1, maxc + 1):
        v = ws.cell(row=HEAD_ROW2, column=c).value
        if v is not None and normalize_text(v) == normalize_text(DEFAULT_COUNT_HEADER):
            return c
    return None


# ----------------------------
# Main generation
# ----------------------------

def generate_mpdt_for_asset(
    template_path: Path,
    output_path: Path,
    lodm_df: pd.DataFrame,
    join2_df: pd.DataFrame,
    data_rows: List[dict],
    join2_key_col: str = "UAID_3",
    class_code_keys: Sequence[str] = ("AssetHierarchyCategory", "HS2_Class", "ClassCode"),
    au_start_col_idx: Optional[int] = None,
    count_col_idx: Optional[int] = None,
    first15: int = FIRST15_PERMANENT,
    logger: Optional[logging.Logger] = None,
) -> None:
    """
    Writes an MPDT workbook cloned from template_path to output_path with updated AU headers and data.
    data_rows: list of dicts for each MPDT row (UAID_2, UAID_3, names, class code, coords, etc.)
    """
    log = logger or logging.getLogger("mpdt_generator")

    # Prepare AU headers from att_matrix + LoDM
    au_row1, au_row2 = build_final_au_headers(template_path, lodm_df, first15)

    # Load template (keep VBA)
    wb = load_workbook(template_path, keep_vba=True)
    # Choose the main MPDT sheet (adapt the name to your template)
    sheet_name = "MPDT Element of Asset" if "MPDT Element of Asset" in wb.sheetnames else wb.sheetnames[0]
    ws = wb[sheet_name]

    # Locate AU start and count columns
    if not au_start_col_idx:
        au_start_col_idx = find_au_start(ws) or col_letter_to_index(DEFAULT_AU_LETTER)
    if not count_col_idx:
        count_col_idx = find_count_col_idx(ws) or (au_start_col_idx + 1)  # fallback near AU

    # Write headers at AU+
    write_au_headers(ws, au_start_col_idx, au_row1, au_row2)

    # Ensure join2 is wide
    join2_wide, detected_key_col = ensure_join2_wide(join2_df)
    key_col = join2_key_col if join2_key_col in (join2_wide.columns if not join2_wide.empty else []) else (detected_key_col or join2_key_col)

    # Write data rows
    row_idx = DATA_FIRST_ROW
    for row in data_rows:
        # Caller should have already filled fixed A..(AU-1) cells. Here we only fill AU+ and count.
        # If you need to write governance columns too, add that logic here.
        # Fill AU+ from join2
        key_val = str(row.get(key_col, "")) if key_col else ""
        fill_au_values_row(ws, row_idx, au_start_col_idx, au_row1, join2_wide, key_col, key_val)

        # Checkerboard and count
        class_code = ""
        for k in class_code_keys:
            v = row.get(k)
            if not is_empty(v):
                class_code = str(v).strip()
                break

        apply_checkerboard(ws, row_idx, au_start_col_idx, au_row1, class_code, lodm_df, first15=first15)
        write_count(ws, row_idx, count_col_idx, lodm_df, class_code)

        row_idx += 1

    # Save result
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    log.info("MPDT written: %s", output_path)


# ----------------------------
# CLI
# ----------------------------

def parse_args():
    ap = argparse.ArgumentParser(description="Generate MPDT per UAID_2 from template att_matrix + LoDM + join2 wide.")
    ap.add_argument("--template", required=True, help="Template MPDT XLSM with att_matrix sheet")
    ap.add_argument("--lodm", required=True, help="LoDM file (CSV/XLSX)")
    ap.add_argument("--join2", required=True, help="SmartForms L3 (wide or normalized) data (CSV/XLSX)")
    ap.add_argument("--data", required=True, help="CSV/XLSX of MPDT rows (UAID_2/UAID_3/ClassCode/coords/...); one row per output row")
    ap.add_argument("--out", required=True, help="Output MPDT path (XLSM/XLSX)")
    ap.add_argument("--join2-key", default="UAID_3", help="Key column in join2 wide used to fetch values (default UAID_3)")
    ap.add_argument("--au-start", help="Override AU start column (letter), e.g., AU")
    ap.add_argument("--count-col", help="Override count column (letter), e.g., AV")
    ap.add_argument("-v", "--verbose", action="count", default=1)
    return ap.parse_args()

def main():
    args = parse_args()
    level = logging.WARNING if args.verbose == 0 else (logging.INFO if args.verbose == 1 else logging.DEBUG)
    logging.basicConfig(level=level, format="%(asctime)s | %(levelname)-8s | %(name)s | %(message)s", datefmt="%Y-%m-%d %H:%M:%S")
    log = logging.getLogger("mpdt_generator_custom")

    template_path = Path(args.template)
    lodm_df = read_table_any(Path(args.lodm))
    join2_df = read_table_any(Path(args.join2))
    data_df = read_table_any(Path(args.data))

    if data_df.empty:
        raise RuntimeError("No data rows provided in --data")
    data_rows = data_df.fillna("").astype(str).to_dict(orient="records")

    au_start_idx = col_letter_to_index(args.au_start) if args.au_start else None
    count_col_idx = col_letter_to_index(args.count_col) if args.count_col else None

    generate_mpdt_for_asset(
        template_path=template_path,
        output_path=Path(args.out),
        lodm_df=lodm_df,
        join2_df=join2_df,
        data_rows=data_rows,
        join2_key_col=args.join2_key,
        au_start_col_idx=au_start_idx,
        count_col_idx=count_col_idx,
        logger=log,
    )

if __name__ == "__main__":
    main()