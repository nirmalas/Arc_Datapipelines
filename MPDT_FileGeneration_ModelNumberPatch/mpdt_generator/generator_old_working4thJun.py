"""
mpdt_generator/generator.py — Generate MPDT (.xlsm) files for target UAID_2 assets.

Mirrors the logic in mpdt_creation_v2.ipynb cell "create_mpdt_output_files()" (code block 6):
  - join1  = AssetsScope2 (DB) ∩ AssetRegisterLevel2 (SmartForms L2) on UAID_2
  - join2  = AssetRegisterLevel3 (SmartForms L3) ∩ AssetsScope3 (DB)  on UAID_3/UAID_2
  - One output MPDT row per join2 row (= one per L3 child asset)
  - Pass 1: mapping_dict → join2 direct → join1 direct (pre-AU columns only)
  - Pass 2: special overrides (ModelContainerID, Discipline, Chainage, Layer …)
  - Pass 3: att_matrix applicability; sets AV column count

Can be run standalone:
  python -m mpdt_generator.generator --workspace . --target-uaid2 HS2-000001416
"""
from __future__ import annotations

import argparse
import difflib
import logging
import copy
import re
import shutil
import subprocess
from pathlib import Path
from typing import Any

import numpy as np
import pandas as pd

from mpdt_generator.midp_resolver import build_model_container_resolver

from utils.common import (
    get_available_path,
    load_config,
    normalize_text,
    resolve_workspace,
    sanitize_filename,
    setup_logger,
    timestamped_dir,
    write_json,
    read_json,
)

try:
    from openpyxl import load_workbook
    from openpyxl.utils import column_index_from_string
    from openpyxl.styles import PatternFill
except ImportError:
    load_workbook = None  # type: ignore
    column_index_from_string = None  # type: ignore
    PatternFill = None  # type: ignore


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _is_empty(val: Any) -> bool:
    if val is None:
        return True
    try:
        if isinstance(val, float) and np.isnan(val):
            return True
    except Exception:
        pass
    return str(val).strip() in ("", "nan", "None", "NaN", "NaT")


def is_applicable(att_tst: Any, asset_type: Any) -> bool:
    """
    Mirrors the notebook VBA logic exactly:
    Returns True if asset_type appears in the att_tst token list or as substring.
    """
    if _is_empty(att_tst) or str(att_tst).strip().upper() in ("ALL", "Y", "YES"):
        return True
    if _is_empty(asset_type):
        return False
    tokens = [t.strip().upper() for t in str(att_tst).replace(",", " ").split() if t.strip()]
    cc_upper = str(asset_type).upper()
    return all(t in cc_upper for t in tokens)


def get_mapped_value(expr: str, join2_row: pd.Series, join1_row: pd.Series | None, deliverable_name: str | None = None) -> Any:
    """Resolve mapping expression like 'join2[Uaid]' or 'join1[AssetName]'."""
    if _is_empty(expr):
        return None

    def _extract(row: pd.Series | None, colname: str, preferred_level: int | None = None) -> Any:
        """Look up colname in row, tolerating join-merge suffixes (_l3, _sform3, etc.).

        preferred_level (2 or 3): when colname is a bare name like 'AssetName' and the
        row contains both 'AssetName_2_l3' and 'AssetName_3_l3', this selects the right
        one.  Pass 3 for join2 (L3) and 2 for join1 (L2).
        """
        if row is None or row.empty:
            return None
        norm_target = normalize_text(colname)

        # 1) exact normalized match
        for col in row.index:
            if normalize_text(col) == norm_target:
                val = row[col]
                return None if _is_empty(val) else val

        # 2) tolerate source suffixes like _sform3/_l3 added during joins
        for col in row.index:
            nc = normalize_text(col)
            if nc.startswith(norm_target + " "):
                val = row[col]
                return None if _is_empty(val) else val

        # 3) check by stripping common join-merge suffix tokens from raw column names
        #    e.g. 'HS2_Class_l3' -> 'HS2_Class' matches target 'HS2_Class'
        suffixes = ("_sform2", "_sform3", "_l2", "_l3")
        for col in row.index:
            raw = str(col)
            base = raw
            for suf in suffixes:
                if raw.lower().endswith(suf):
                    base = raw[: -len(suf)]
                    break
            if normalize_text(base) == norm_target:
                val = row[col]
                return None if _is_empty(val) else val

        # 4) bare name like 'AssetName' should resolve to 'AssetName_3_l3' (L3) or
        #    'AssetName_2_l2' (L2).  Try preferred_level first, then other levels.
        #    The level digit is preserved in the matched column name — it is NOT stripped.
        _level_sfx = re.compile(r"_(\d+)$")
        for level in ([preferred_level] if preferred_level else []) + [3, 2, 1, 4]:
            suffix_norm = f"{norm_target} {level}"   # e.g. "assetname 3"
            for col in row.index:
                raw = str(col)
                base = raw
                for suf in suffixes:
                    if raw.lower().endswith(suf):
                        base = raw[: -len(suf)]
                        break
                if normalize_text(base) == suffix_norm:
                    val = row[col]
                    return None if _is_empty(val) else val
        return None

    m = re.match(r"join([12])\[(.+?)\]", str(expr).strip(), re.IGNORECASE)
    if m:
        source, colname = m.group(1), m.group(2)
        colname = str(colname).strip().strip("'\"")
        primary = join2_row if source == "2" else (join1_row if join1_row is not None else pd.Series(dtype=object))
        fallback = join1_row if source == "2" else join2_row
        # join2 -> L3 level 3, join1 -> L2 level 2
        pref = 3 if source == "2" else 2
        val = _extract(primary, colname, preferred_level=pref)
        if _is_empty(val):
            val = _extract(fallback, colname, preferred_level=pref)
        return val

    # Support dataframe-style mapping expressions from the notebook mapping file,
    # e.g. asset_class_df['AssetDescription_3'] or join2_df["UAID_3"].
    m2 = re.match(
        r"(?:asset_class_df|join1_df|join2_df|join1|join2)\s*\[\s*['\"]([^'\"]*)['\"]\s*\]",
        str(expr).strip(),
        re.IGNORECASE,
    )
    if m2:
        colname = m2.group(1).strip()
        if not colname:
            return None
        val = _extract(join2_row, colname, preferred_level=3)
        if _is_empty(val):
            val = _extract(join1_row, colname, preferred_level=2)
        return val

    # Special-case tokens that should map to the output/deliverable name
    lit = str(expr).strip()
    if lit and lit.lower() in ("deliverablename", "deliverable_name", "deliverable", "filename", "outputfilename", "output_file"):
        return deliverable_name

    # Notebook mapping file contains instruction literals for round-2 derived fields.
    # Return None here so round-2 logic can compute the final values.
    lit_norm = normalize_text(lit)
    if lit_norm in (
        "3rd part of model file name",
        "concatanation of ag ak joined with",
        "existing mpdt if uaid not present in existing mpdt then leave blank",
    ):
        return None

    return expr  # Literal


def _mapping_expr_for_column(mapping_dict: dict[str, str], column_name: str) -> str:
    """Resolve mapping expression with normalized-key fallback (handles trailing spaces)."""
    expr = mapping_dict.get(column_name)
    if expr is not None:
        return expr
    ncol = normalize_text(column_name)
    for k, v in mapping_dict.items():
        if normalize_text(k) == ncol:
            return v
    return ""


# ---------------------------------------------------------------------------
# Data join builders
# ---------------------------------------------------------------------------

def _uaid_col(df: pd.DataFrame, candidates: list[str]) -> str | None:
    for c in candidates:
        nc = normalize_text(c)
        for col in df.columns:
            if normalize_text(col) == nc:
                return col
    return None



def _upper_key(value: Any) -> str:
    """Return a normalized uppercase lookup key for UAID-style values."""
    if _is_empty(value):
        return ""
    return str(value).strip().upper()


def _first_row_by_upper_key(df: "pd.DataFrame | None", key_col: str | None) -> dict[str, pd.Series]:
    """Index a DataFrame by a normalized key column, keeping the first row for each key.

    This is used by generate_mpdt_batch() to avoid repeatedly filtering large
    Scope/SmartForms DataFrames for every generated MPDT file. Empty keys are
    ignored and duplicate keys keep the first non-empty occurrence, matching the
    previous .iloc[0] behaviour.
    """
    if df is None or df.empty or not key_col or key_col not in df.columns:
        return {}

    out: dict[str, pd.Series] = {}
    for _, row in df.iterrows():
        key = _upper_key(row.get(key_col))
        if key and key not in out:
            out[key] = row
    return out


def _group_df_by_upper_key(df: "pd.DataFrame | None", key_col: str | None) -> dict[str, pd.DataFrame]:
    """Index a DataFrame by a normalized key column, keeping all rows per key."""
    if df is None or df.empty or not key_col or key_col not in df.columns:
        return {}

    work = df.copy()
    work["__mpdt_lookup_key__"] = work[key_col].map(_upper_key)
    work = work[work["__mpdt_lookup_key__"] != ""]
    if work.empty:
        return {}

    groups: dict[str, pd.DataFrame] = {}
    for key, group in work.groupby("__mpdt_lookup_key__", sort=False):
        groups[str(key)] = group.drop(columns=["__mpdt_lookup_key__"])
    return groups


def build_joins(
    scope2: pd.DataFrame,
    scope3: pd.DataFrame,
    sf_l2: pd.DataFrame,
    sf_l3: pd.DataFrame,
    uaid2: str,
    logger: logging.Logger,
) -> tuple[pd.Series, pd.Series]:
    """
    join1 = AssetsScope2 ∩ SmartForms_L2 on UAID_2
    join2 = AssetsScope3 ∩ SmartForms_L3 on UAID_3 (filtered by parent UAID_2)
    Returns (join1_row, join2_row) as pd.Series.
    """
    join1_row = pd.Series(dtype=object)
    join2_row = pd.Series(dtype=object)

    # join1: scope2 + sf_l2
    if not scope2.empty and not sf_l2.empty:
        uaid2_col_s2 = _uaid_col(scope2, ["UAID_2", "Uaid_2", "Uaid", "UAID"])
        uaid2_col_sf = _uaid_col(sf_l2, ["UAID_2", "Uaid_2", "Uaid", "UAID", "Asset_ID"])
        if uaid2_col_s2 and uaid2_col_sf:
            uaid2_key = str(uaid2).strip().upper()
            s2_row = scope2[scope2[uaid2_col_s2].fillna("").astype(str).str.strip().str.upper() == uaid2_key]
            sf2_row = sf_l2[sf_l2[uaid2_col_sf].fillna("").astype(str).str.strip().str.upper() == uaid2_key]
            if not s2_row.empty and not sf2_row.empty:
                merged = pd.merge(
                    s2_row.add_suffix("_l3"), sf2_row.add_suffix("_sform2"),
                    left_index=True, right_index=True,
                )
                if not merged.empty:
                    join1_row = merged.iloc[0]
            elif not s2_row.empty:
                join1_row = s2_row.iloc[0]
            elif not sf2_row.empty:
                join1_row = sf2_row.iloc[0]

    # join2: scope3 + sf_l3 (filtered by parent UAID_2)
    if not scope3.empty and not sf_l3.empty:
        uaid2_col_s3 = _uaid_col(scope3, ["UAID_2", "Uaid_2", "ParentUaid", "Parent_UAID"])
        uaid3_col_s3 = _uaid_col(scope3, ["UAID_3", "Uaid_3", "Uaid", "UAID"])
        uaid3_col_sf = _uaid_col(sf_l3, ["UAID_3", "Uaid_3", "Uaid", "UAID", "Asset_ID"])
        if uaid2_col_s3 and uaid3_col_s3 and uaid3_col_sf:
            uaid2_key = str(uaid2).strip().upper()
            s3_for_parent = scope3[
                scope3[uaid2_col_s3].fillna("").astype(str).str.strip().str.upper() == uaid2_key
            ]
            if not s3_for_parent.empty:
                s3_uaid3s = set(
                    s3_for_parent[uaid3_col_s3]
                    .fillna("")
                    .astype(str)
                    .str.strip()
                    .str.upper()
                    .tolist()
                )
                sf3_keys = sf_l3[uaid3_col_sf].fillna("").astype(str).str.strip().str.upper()
                sf3_rows = sf_l3[sf3_keys.isin(s3_uaid3s)]
                if not sf3_rows.empty:
                    # Join on normalized UAID_3 values to avoid case/spacing mismatch.
                    left = sf3_rows.add_suffix("_sform3").copy()
                    right = s3_for_parent.add_suffix("_l3").copy()
                    left_key = f"{uaid3_col_sf}_sform3"
                    right_key = f"{uaid3_col_s3}_l3"
                    left["__join_key__"] = left[left_key].fillna("").astype(str).str.strip().str.upper()
                    right["__join_key__"] = right[right_key].fillna("").astype(str).str.strip().str.upper()
                    merged = pd.merge(
                        left,
                        right,
                        on="__join_key__",
                        how="inner",
                    )
                    if not merged.empty:
                        join2_row = merged.iloc[0]
                else:
                    join2_row = s3_for_parent.iloc[0]

    return join1_row, join2_row


# ---------------------------------------------------------------------------
# Pass 1 & 2: populate a single row
# ---------------------------------------------------------------------------

def populate_row(
    columns: list[str],
    mapping_dict: dict[str, str],
    join2_row: pd.Series,
    join1_row: pd.Series,
    uaid2: str,
    deliverable_name: str | None = None,
    model_container_id: str | None = None,
) -> dict[str, Any]:
    row: dict[str, Any] = {}

    print(f"Populating row for UAID_2={uaid2}, deliverable={deliverable_name}, columns={columns}, mapping_dict={mapping_dict}")
    
    # Pass 1 — mapping
    for col in columns:
        expr = _mapping_expr_for_column(mapping_dict, col)
        val = get_mapped_value(expr or "", join2_row, join1_row, deliverable_name) if expr else None
        row[col] = val

    # Pass 2 — special overrides
    def _find(*candidates: str) -> Any:
        for c in candidates:
            val = get_mapped_value(f"join2[{c}]", join2_row, None, deliverable_name)
            if not _is_empty(val):
                return val
            val = get_mapped_value(f"join1[{c}]", pd.Series(dtype=object), join1_row, deliverable_name)
            if not _is_empty(val):
                return val
        return None

    for col in columns:
        nc = normalize_text(col)
        if "model container" in nc or "container id" in nc:
            row[col] = model_container_id or None
        elif nc == "discipline":
            row[col] = _find("Discipline", "DisciplineCode")
        elif "chainage" in nc:
            row[col] = _find("Chainage", "StartChainage", "Start_Chainage")
        elif nc in ("description", "asset description"):
            row[col] = _find("Level Description", "Description", "AssetDescription", "Asset_Description")
        elif nc in ("uaid 2", "uaid_2", "uaid2", "level 2 asset id"):
            row[col] = uaid2

    # Explicitly ensure common L3 fields come from join2 (L3) when present
    # AssetName_3 and AssetDescription_3 should come from the L3 assetscope3 table
    for hdr in ("AssetName_3", "AssetDescription_3"):
        if hdr in columns:
            val = get_mapped_value(f"join2[{hdr.split('_')[0]}]", join2_row, join1_row, deliverable_name)
            if not _is_empty(val):
                row[hdr] = val

    return row


# ---------------------------------------------------------------------------
# Round 2: compute derived MPDT columns from other MPDT columns + external data
# ---------------------------------------------------------------------------

_BLANK_LITERALS = {"blank", "nan", "none", "nat", ""}


def _apply_round2(
    df: pd.DataFrame,
    columns: list[str],
    uaid2: str,
    deliverable_name: str,
    software_model_part_id_value: str,
    lodm_df: pd.DataFrame,
    control_df: pd.DataFrame,
    pw_df: pd.DataFrame,
) -> pd.DataFrame:
    """
    Round 2 post-processing on the fully-built MPDT DataFrame.

    Computes columns that depend on other MPDT column values or external lookups:
      - 'blank'/'Blank' literals → None (empty cell in output)
      - Discipline         = 3rd segment of deliverable_name split by '-'
      - AssetDescription_3 = LoDM ClassName where ClassCode = AssetHierarchyCategory
      - Layer              = Discipline-Classification-Presentation_Description (skip empty)
      - '  ' (count)      = number of LoDM attribute rows for the row's ClassCode
    - Software Model Part ID no. = value read from latest existing MPDT doc for UAID_2
    """
    # ------------------------------------------------------------------ #
    # 0. Normalise 'blank' / 'Blank' literal → None                       #
    # ------------------------------------------------------------------ #
    for col in df.columns:
        df[col] = df[col].apply(
            lambda v: None
            if v is not None and str(v).strip().lower() in _BLANK_LITERALS
            else v
        )

    # ------------------------------------------------------------------ #
    # 1. Discipline = 3rd segment of deliverable_name (split on '-')       #
    # ------------------------------------------------------------------ #
    discipline = ""
    if deliverable_name:
        parts = deliverable_name.split("-")
        discipline = parts[2].strip() if len(parts) > 2 else ""
    disc_col = next((c for c in columns if normalize_text(c) == "discipline"), None)
    if disc_col and disc_col in df.columns and discipline:
        # Only override when we have a real parsed value.
        # For fallback names like MPDT_HS2-..., keep whatever pass-1 populated.
        df[disc_col] = discipline

    # ------------------------------------------------------------------ #
    # 2. AssetDescription_3 = LoDM ClassName for the row's ClassCode       #
    # ------------------------------------------------------------------ #
    # The ClassCode lives in AssetHierarchyCategory (=HS2_Class from scope3)
    hier_col = next(
        (c for c in df.columns if normalize_text(c) in
         ("assethierarchycategory", "hs2_class", "classcode")),
        None,
    )
    if "AssetDescription_3" in df.columns and not lodm_df.empty and hier_col:
        cc_col = next(
            (c for c in lodm_df.columns if normalize_text(c) == "classcode"), None
        )
        cn_col = next(
            (c for c in lodm_df.columns
             if normalize_text(c) in ("classname", "class name")),
            None,
        )
        if cc_col and cn_col:
            name_map = (
                lodm_df[[cc_col, cn_col]]
                .drop_duplicates(subset=[cc_col])
                .set_index(cc_col)[cn_col]
                .to_dict()
            )
            df["AssetDescription_3"] = df[hier_col].apply(
                lambda code: name_map.get(str(code).strip())
                if not _is_empty(code) else None
            )

    # ------------------------------------------------------------------ #
    # 3. Layer = concat of [Discipline, Classification, Presentation,      #
    #           Description, View] with '-', where Presentation+Description#
    #           are joined by '_' to match the established MPDT format     #
    # ------------------------------------------------------------------ #
    class_col = next((c for c in columns if normalize_text(c) == "classification"), None)

    # Fill Description from control file (by ClassCode) and then AssetDescription_3 fallback.
    desc_col = next((c for c in columns if normalize_text(c) == "description"), None)
    control_desc_map = {}
    control_disc_map = {}
    if not control_df.empty:
        code_col = next((c for c in control_df.columns if normalize_text(c) == "code"), None)
        lvl_desc_col = next((c for c in control_df.columns if normalize_text(c) == "level description"), None)
        lvl_disc_col = next((c for c in control_df.columns if normalize_text(c) == "discipline"), None)
        if code_col and lvl_desc_col:
            tmp = control_df[[code_col, lvl_desc_col]].dropna().drop_duplicates(subset=[code_col])
            control_desc_map = {str(k).strip(): str(v).strip() for k, v in zip(tmp[code_col], tmp[lvl_desc_col])}
        if code_col and lvl_disc_col:
            tmp2 = control_df[[code_col, lvl_disc_col]].dropna().drop_duplicates(subset=[code_col])
            control_disc_map = {str(k).strip(): str(v).strip() for k, v in zip(tmp2[code_col], tmp2[lvl_disc_col])}

    if desc_col and desc_col in df.columns:
        def _desc_fix(row: pd.Series):
            cur = row.get(desc_col)
            if _is_empty(cur) or normalize_text(str(cur)) == normalize_text("control_files_df['Level Description']"):
                code = ""
                if class_col and not _is_empty(row.get(class_col)):
                    code = str(row.get(class_col)).strip()
                elif hier_col and not _is_empty(row.get(hier_col)):
                    code = str(row.get(hier_col)).strip()
                if code and code in control_desc_map:
                    return control_desc_map[code]
                if "AssetDescription_3" in row.index:
                    return row.get("AssetDescription_3")
            if not _is_empty(cur):
                code = ""
                if class_col and not _is_empty(row.get(class_col)):
                    code = str(row.get(class_col)).strip()
                elif hier_col and not _is_empty(row.get(hier_col)):
                    code = str(row.get(hier_col)).strip()
                if code and code in control_desc_map:
                    return control_desc_map[code]
            return cur
        df[desc_col] = df.apply(_desc_fix, axis=1)

    # Fill Discipline from control file mapping when empty.
    if disc_col and disc_col in df.columns:
        def _disc_fix(row: pd.Series):
            cur = row.get(disc_col)
            if _is_empty(cur):
                code = ""
                if class_col and not _is_empty(row.get(class_col)):
                    code = str(row.get(class_col)).strip()
                elif hier_col and not _is_empty(row.get(hier_col)):
                    code = str(row.get(hier_col)).strip()
                if code and code in control_disc_map:
                    return control_disc_map[code]
                if code:
                    if "_" in code:
                        return code.split("_", 1)[0]
                    if "-" in code:
                        return code.split("-", 1)[0]
            return cur
        df[disc_col] = df.apply(_disc_fix, axis=1)

    layer_col = next((c for c in columns if normalize_text(c) == "layer"), None)
    ag_ak_names = ["Discipline", "Classification", "Presentation", "Description", "View"]
    if layer_col and layer_col in df.columns:
        present_layer = [c for c in ag_ak_names if c in df.columns]

        def _build_layer(row: pd.Series) -> str | None:
            segments: list[str] = []
            pres_val = str(row["Presentation"]).strip() if "Presentation" in row and not _is_empty(row["Presentation"]) else ""
            desc_val = str(row["Description"]).strip() if "Description" in row and not _is_empty(row["Description"]) else ""
            for col_name in present_layer:
                v = row.get(col_name)
                if _is_empty(v):
                    continue
                sv = str(v).strip()
                if col_name == "Presentation":
                    # Will be fused with Description using '_'
                    if desc_val:
                        segments.append(f"{sv}_{desc_val}")
                    else:
                        segments.append(sv)
                elif col_name == "Description":
                    # Already handled above with Presentation
                    if not pres_val:
                        segments.append(sv)
                else:
                    segments.append(sv)
            result = "-".join(segments)
            return result if result else None

        df[layer_col] = df[present_layer].apply(_build_layer, axis=1)

    # ------------------------------------------------------------------ #
    # 4. Complete/Count columns from LoDM attribute counts                  #
    # ------------------------------------------------------------------ #
    complete_col = next((c for c in columns if normalize_text(c) == "complete"), None)
    if complete_col is None:
        complete_col = next((c for c in columns if str(c).strip() == ""), None)

    count_col = next((c for c in columns if normalize_text(c) == "count"), None)
    if count_col is None:
        # Template AV column is pandas-renamed from '  ' to '  .1' (whitespace-only + .N suffix).
        # Match only columns whose entire string is whitespace + optional .digits (e.g. '  .1').
        count_col = next(
            (c for c in columns if re.fullmatch(r"\s+\.\d+", str(c))),
            None,
        )
    if count_col is None and complete_col and complete_col in columns:
        i = columns.index(complete_col)
        if i + 1 < len(columns) and columns[i + 1] != complete_col:
            count_col = columns[i + 1]

    if complete_col and complete_col in df.columns:
        df[complete_col] = None

    if count_col and count_col in df.columns and not lodm_df.empty and (hier_col or class_col):
        cc_col2 = next(
            (c for c in lodm_df.columns if normalize_text(c) == "classcode"), None
        )
        if cc_col2:
            count_map = lodm_df[cc_col2].value_counts().to_dict()
            def _row_count(row: pd.Series):
                candidates = []
                if class_col and not _is_empty(row.get(class_col)):
                    candidates.append(str(row.get(class_col)).strip())
                if hier_col and not _is_empty(row.get(hier_col)):
                    candidates.append(str(row.get(hier_col)).strip())
                for code in candidates:
                    if code in count_map:
                        return count_map.get(code)
                return None
            df[count_col] = df.apply(_row_count, axis=1)

    # ------------------------------------------------------------------ #
    # 4b. Fill 'strAttr HS2 Asset Classification' from AssetHierarchyCategory  #
    # ------------------------------------------------------------------ #
    hs2_cls_attr = next(
        (c for c in columns if normalize_text(c) == "strattr hs2 asset classification"),
        None,
    )
    if hs2_cls_attr and hs2_cls_attr in df.columns and hier_col and hier_col in df.columns:
        mask = df[hs2_cls_attr].apply(_is_empty)
        df.loc[mask, hs2_cls_attr] = df.loc[mask, hier_col]

    # ------------------------------------------------------------------ #
    # 5. Software Model Part ID no. = latest existing MPDT document name    #
    # ------------------------------------------------------------------ #
    sw_col = next((c for c in columns if normalize_text(c) == "software model part id no"), None)
    if sw_col is None:
        sw_col = next((c for c in columns if str(c).strip() == "Unnamed: 43"), None)
    if sw_col and sw_col in df.columns:
        df[sw_col] = software_model_part_id_value if software_model_part_id_value else None

    # 6. OSGB Easting/Northing fallback fill from alternate matching columns
    easting_targets = [c for c in columns if normalize_text(c) in ("osgb easting", "numattr osgb easting")]
    northing_targets = [c for c in columns if normalize_text(c) in ("osgb northing", "numattr osgb northing")]

    def _first_non_empty(row: pd.Series, norm_name: str):
        for c in row.index:
            if normalize_text(c) == norm_name and not _is_empty(row.get(c)):
                return row.get(c)
        return None

    for target in easting_targets:
        if target in df.columns:
            df[target] = df.apply(
                lambda r: _first_non_empty(r, "numattr osgb easting")
                if _is_empty(r.get(target)) else r.get(target),
                axis=1,
            )
    for target in northing_targets:
        if target in df.columns:
            df[target] = df.apply(
                lambda r: _first_non_empty(r, "numattr osgb northing")
                if _is_empty(r.get(target)) else r.get(target),
                axis=1,
            )

    return df


def _latest_existing_pw_mpdt_name(uaid2: str, pw_df: pd.DataFrame) -> str:
    """Return latest existing MPDT document name for this UAID_2 from PW extract, else ''."""
    if pw_df is None or pw_df.empty:
        return ""

    candidates = [c for c in ("PW_UAID", "ASSET_ID", "UAID_2", "ZZ_LBL_UAIDL2", "ZZ_LBL_UAIDL1") if c in pw_df.columns]
    if not candidates:
        return ""

    uaid_key = str(uaid2).strip().upper()
    subset = pd.DataFrame()
    for c in candidates:
        hit = pw_df[pw_df[c].fillna("").astype(str).str.strip().str.upper() == uaid_key]
        if not hit.empty:
            subset = pd.concat([subset, hit], ignore_index=True)
    if subset.empty:
        return ""

    file_col = "FileName" if "FileName" in subset.columns else ("DocumentName" if "DocumentName" in subset.columns else None)
    if file_col is None:
        return ""

    names = subset[file_col].fillna("").astype(str).str.lower()
    subset = subset[names.str.endswith((".xlsm", ".xlsx", ".xls"))].copy()
    if subset.empty:
        return ""

    if "FileUpdated_parsed" in subset.columns:
        subset["__sort_date__"] = pd.to_datetime(subset["FileUpdated_parsed"], errors="coerce")
    elif "FileUpdated" in subset.columns:
        subset["__sort_date__"] = pd.to_datetime(subset["FileUpdated"], errors="coerce")
    else:
        subset["__sort_date__"] = pd.NaT

    subset = subset.sort_values("__sort_date__", ascending=False, na_position="last")
    best = subset.iloc[0]

    doc_name = str(best.get("DocumentName", "")).strip()
    if doc_name:
        return doc_name
    fn = str(best.get("FileName", "")).strip()
    return Path(fn).stem if fn else ""


def _latest_existing_pw_mpdt_row(uaid2: str, pw_df: pd.DataFrame) -> pd.Series | None:
    """Return the latest existing MPDT row for this UAID_2 from PW extract, else None."""
    if pw_df is None or pw_df.empty:
        return None

    candidates = [c for c in ("PW_UAID", "ASSET_ID", "UAID_2", "ZZ_LBL_UAIDL2", "ZZ_LBL_UAIDL1") if c in pw_df.columns]
    if not candidates:
        return None

    uaid_key = str(uaid2).strip().upper()
    subset = pd.DataFrame()
    for c in candidates:
        hit = pw_df[pw_df[c].fillna("").astype(str).str.strip().str.upper() == uaid_key]
        if not hit.empty:
            subset = pd.concat([subset, hit], ignore_index=True)
    if subset.empty:
        return None

    file_col = "FileName" if "FileName" in subset.columns else ("DocumentName" if "DocumentName" in subset.columns else None)
    if file_col is None:
        return None

    names = subset[file_col].fillna("").astype(str).str.lower()
    subset = subset[names.str.endswith((".xlsm", ".xlsx", ".xls"))].copy()
    if subset.empty:
        return None

    if "FileUpdated_parsed" in subset.columns:
        subset["__sort_date__"] = pd.to_datetime(subset["FileUpdated_parsed"], errors="coerce")
    elif "FileUpdated" in subset.columns:
        subset["__sort_date__"] = pd.to_datetime(subset["FileUpdated"], errors="coerce")
    else:
        subset["__sort_date__"] = pd.NaT

    subset = subset.sort_values("__sort_date__", ascending=False, na_position="last")
    return subset.iloc[0]


def _resolve_existing_mpdt_path(workspace: Path, pw_row: pd.Series) -> Path | None:
    """Best-effort resolve of a locally accessible MPDT file path for a PW row."""
    raw_paths = [
        str(pw_row.get("FullPath", "")).strip(),
        str(pw_row.get("FileName", "")).strip(),
        str(pw_row.get("DocumentName", "")).strip(),
    ]

    for raw in raw_paths:
        if not raw:
            continue
        p = Path(raw)
        if p.is_absolute() and p.exists():
            return p
        p2 = (workspace / raw).resolve()
        if p2.exists():
            return p2

    # Fall back to a local name search under Input/ and Output/.
    filename_candidates = [
        str(pw_row.get("FileName", "")).strip(),
        str(pw_row.get("DocumentName", "")).strip(),
    ]
    stems = {Path(v).name for v in filename_candidates if v}
    for root_rel in ("Input", "Output"):
        root = (workspace / root_rel).resolve()
        if not root.exists():
            continue
        for s in stems:
            if not s:
                continue
            for hit in root.rglob(s):
                if hit.is_file():
                    return hit
    return None


def _download_existing_mpdt_from_projectwise(
    workspace: Path,
    cfg: dict,
    document_name: str,
    logger: logging.Logger,
) -> Path | None:
    """Download latest MPDT document from ProjectWise using a configured PowerShell script."""
    if not document_name:
        return None

    pw_cfg = cfg.get("pw", {})
    script_rel = pw_cfg.get("ps1_download", "Scripts/PWPS_Download_MPDT.ps1")
    script = (workspace / script_rel).resolve()
    if not script.exists():
        logger.warning("  PW download script not found: %s", script)
        return None

    out_dir = (workspace / "Output" / "PW_DownloadCache").resolve()
    out_dir.mkdir(parents=True, exist_ok=True)

    cmd = [
        "pwsh",
        "-File",
        str(script),
        "-DocumentName",
        str(document_name),
        "-OutputDir",
        str(out_dir),
        "-DatasourceName",
        str(pw_cfg.get("datasource", "arcadis-uk-pw.bentley.com:arcadis-uk-07")),
        "-UserName",
        str(pw_cfg.get("username", "_asc_user_automation")),
    ]

    try:
        # Keep interactive terminal behavior for PW password prompt.
        proc = subprocess.run(cmd, capture_output=True, text=True, timeout=600)
        if proc.returncode != 0:
            logger.warning("  PW download failed for '%s': %s", document_name, (proc.stderr or proc.stdout)[-1200:])
            return None

        combined = f"{proc.stdout}\n{proc.stderr}"
        m = re.search(r"DownloadedFile\s*=\s*(.+)", combined)
        if m:
            p = Path(m.group(1).strip().strip('"'))
            if p.exists():
                logger.info("  Downloaded existing MPDT from PW: %s", p.name)
                return p

        # Fallback: newest matching file in cache directory.
        stem = Path(document_name).stem.lower()
        hits = [p for p in out_dir.glob("*.xls*") if stem in p.stem.lower()]
        if hits:
            hits.sort(key=lambda p: p.stat().st_mtime, reverse=True)
            logger.info("  Downloaded existing MPDT from PW: %s", hits[0].name)
            return hits[0]
    except Exception as exc:
        logger.warning("  Could not download '%s' from ProjectWise: %s", document_name, exc)

    return None


def _extract_sw_model_part_id_from_existing_mpdt(
    workspace: Path,
    cfg: dict,
    uaid2: str,
    pw_df: pd.DataFrame,
    columns: list[str],
    logger: logging.Logger,
) -> str:
    """Read Software Model Part ID no. from the latest existing MPDT file for UAID_2."""
    if load_workbook is None:
        return ""

    pw_row = _latest_existing_pw_mpdt_row(uaid2, pw_df)
    if pw_row is None:
        return ""

    mpdt_path = _resolve_existing_mpdt_path(workspace, pw_row)
    if mpdt_path is None:
        # Required behavior: download latest existing MPDT from ProjectWise, then read value.
        doc_name = str(pw_row.get("DocumentName", "")).strip() or str(pw_row.get("FileName", "")).strip()
        mpdt_path = _download_existing_mpdt_from_projectwise(workspace, cfg, doc_name, logger)
    if mpdt_path is None:
        logger.info("  No local existing MPDT file found for %s; Software Model Part ID left blank.", uaid2)
        return ""

    try:
        wb = load_workbook(str(mpdt_path), read_only=True, data_only=True, keep_vba=True)
        ws = wb["MPDT Element of Asset"] if "MPDT Element of Asset" in wb.sheetnames else wb.active
        if ws is None:
            return ""

        # Prefer semantic header match in row 2, fallback to template-positioned column.
        sw_col_idx = None
        for ci in range(1, ws.max_column + 1):
            hv = ws.cell(row=2, column=ci).value
            if normalize_text(str(hv or "")) == "software model part id no":
                sw_col_idx = ci
                break
        if sw_col_idx is None:
            sw_col = next((c for c in columns if normalize_text(c) == "software model part id no"), None)
            if sw_col and sw_col in columns:
                sw_col_idx = columns.index(sw_col) + 1
        if sw_col_idx is None:
            sw_col_idx = 44  # historical template position (AR)

        for ri in range(3, ws.max_row + 1):
            v = ws.cell(row=ri, column=sw_col_idx).value
            if not _is_empty(v):
                return str(v).strip()
    except Exception as exc:
        logger.warning("  Could not read existing MPDT '%s' for Software Model Part ID: %s", mpdt_path, exc)
    return ""


def _norm_code_variants(code: str) -> set[str]:
    s = str(code or "").strip()
    if not s:
        return set()
    return {s.lower(), s.replace("_", "-").lower(), s.replace("-", "_").lower()}


def _norm_attr_text(text: Any) -> str:
    s = normalize_text(str(text or ""))
    s = re.sub(r"^(strattr|numattr|intattr)\s+", "", s).strip()
    s = s.replace("(m)", "").replace("m3", "").replace("m2", "")
    s = re.sub(r"\s+", " ", s).strip()
    return s


def _lodm_attribute_norm_set_for_class(lodm_df: pd.DataFrame, class_code: str) -> set[str]:
    if lodm_df.empty or not class_code:
        return set()
    cc_col = next((c for c in lodm_df.columns if normalize_text(c) == "classcode"), None)
    if not cc_col:
        return set()
    vars_ = _norm_code_variants(class_code)
    if not vars_:
        return set()

    m = lodm_df[lodm_df[cc_col].fillna("").astype(str).str.strip().str.lower().isin(vars_)].copy()
    if m.empty:
        return set()

    cand_cols = [
        c for c in lodm_df.columns
        if normalize_text(c) in ("atttypename", "atttypedescription", "attrtypedisplayname", "attribute", "attributename")
    ]
    out: set[str] = set()
    for c in cand_cols:
        vals = m[c].dropna().astype(str).str.strip().tolist()
        out.update({t for t in (_norm_attr_text(v) for v in vals) if t})
    return out


_RE_ATT_SEP = re.compile(r'[\s_\-\.:]+')


def _norm_att_code(s: str) -> str:
    """Normalise an AttTypeName code for robust comparison.

    Strips separators (space, underscore, hyphen, dot, colon) and lowercases
    so that e.g. 'Com_Dscrptn' (template) matches 'Com:Dscrptn' (LoDM).
    """
    return _RE_ATT_SEP.sub('', str(s)).lower()


def _lodm_atttype_names_for_class(lodm_df: pd.DataFrame, class_code: str) -> set[str]:
    """Return normalised AttTypeName codes for the given ClassCode.

    Returns only the AttTypeName column values (e.g. 'NmnlSrfcAr', 'Mtrl',
    'MtrlPp'), normalised via *_norm_att_code* so that separator differences
    between the template (underscores) and LoDM (colons) are ignored.  Used
    together with Row 1 short codes for unambiguous LoDM matching.
    """
    if lodm_df.empty or not class_code:
        return set()
    cc_col = next((c for c in lodm_df.columns if normalize_text(c) == "classcode"), None)
    att_col = next((c for c in lodm_df.columns if normalize_text(c) == "atttypename"), None)
    if not cc_col or not att_col:
        return set()
    vars_ = _norm_code_variants(class_code)
    rows = lodm_df[lodm_df[cc_col].fillna("").astype(str).str.strip().str.lower().isin(vars_)]
    return {_norm_att_code(v) for v in rows[att_col].dropna() if str(v).strip()}


def _match_lodm_attrs_to_mpdt_columns(attr_norms: set[str], candidate_columns: list[str]) -> set[str]:
    """Map LoDM attributes to MPDT column names by normalized exact/prefix/fuzzy matching."""
    if not attr_norms:
        return set()

    mpdt_norm = {c: _norm_attr_text(c) for c in candidate_columns}
    out: set[str] = set()

    # LoDM attributes that already have an exact MPDT column match.
    # These are excluded from the prefix check to avoid false positives where a
    # more-specific MPDT column (e.g. "material (pipe)") is erroneously included
    # because it starts with a shorter LoDM attr that already has an exact match
    # (e.g. "material").
    mpdt_norms_set = set(mpdt_norm.values())
    exact_matched_attrs = attr_norms & mpdt_norms_set

    for col, nc in mpdt_norm.items():
        if not nc:
            continue
        if nc in attr_norms:
            out.add(col)
            continue
        # Prefix match: handles cases where MPDT column has extra context appended to the
        # LoDM attribute name, e.g. LoDM='type', MPDT='type (layer description text)';
        # or LoDM='net volume', MPDT='net volume (m3)'.
        # Guard: skip attrs that already matched exactly to avoid including a more-specific
        # MPDT column that merely starts with the shorter LoDM attr name.
        for a in attr_norms:
            if a in exact_matched_attrs:
                continue
            if (nc.startswith(a + " ") or nc.startswith(a + "(")
                    or a.startswith(nc + " ") or a.startswith(nc + "(")):
                out.add(col)
                break
        if col in out:
            continue
        # Fuzzy fallback for formatting differences.
        best = 0.0
        for a in attr_norms:
            r = difflib.SequenceMatcher(a=nc, b=a).ratio()
            if r > best:
                best = r
        if best >= 0.82:
            out.add(col)
    return out



# Row-1 AttTypeName/DB-field codes and Row-2 headers that must always remain
# in the generated MPDT, even though they are not per-class LoDM attributes.
# Everything else from AW onwards is treated as a class-specific attribute and
# is physically removed unless the selected ClassCode(s) explicitly require it.
_PERMANENT_MPDT_ROW1_CODES: set[str] = {
    "uaid2",
    "assetname2",
    "assetdescription2",
    "hs2class",
    "startchainage",
    "endchainage",
    "osgbeasting",
    "osgbnorthing",
    "snakegrideasting",
    "snakegridnorthing",
    "uaid1",
    "uniclassid",
    "uaid3",
}

_PERMANENT_MPDT_HEADER_NORMS: set[str] = {
    "strattr scope level 2 uaid",
    "strattr asset name 2",
    "strattr asset description 2",
    "strattr hs2 asset classification",
    "numattr start chainage",
    "numattr end chainage",
    "numattr osgb easting",
    "numattr osgb northing",
    "numattr snake grid easting",
    "numattr snake grid northing",
    "strattr scope level 1 uaid",
    "strattr uniclass 2015",
    "strattr scope level 3 uaid",
}


def _is_permanent_mpdt_attribute_column(col: str, row1_code: str | None = None) -> bool:
    """Return True for AW+ columns that are fixed MPDT metadata fields.

    The previous filter preserved any column whose Row-1 code was not found
    anywhere in the LoDM, assuming it was governance metadata. That is too broad:
    template attribute columns such as Vlm, LwstLvl, NmnlSrfcAr, FFwk, Ftop, and
    EmbdddLngth may be absent from the current LoDM file but still appear in the
    sample template. Those must be removed unless the selected ClassCode(s)
    explicitly include them. Therefore permanent fields are now allow-listed.
    """
    if _norm_att_code(row1_code or "") in _PERMANENT_MPDT_ROW1_CODES:
        return True
    return normalize_text(col) in _PERMANENT_MPDT_HEADER_NORMS

def _build_row_allowed_columns(
    mpdt_df: pd.DataFrame,
    columns: list[str],
    lodm_df: pd.DataFrame,
    au_start_idx: int,
    logger: "logging.Logger | None" = None,
    row1_codes: "dict[str, str] | None" = None,
) -> tuple[list[set[str]], set[str]]:
    """Return per-row and union sets of AU+ MPDT columns allowed by LoDM attributes.

    When *row1_codes* is provided (mapping Row-2 header → Row-1 AttTypeName
    short code), exact AttTypeName matching is used: only columns whose Row 1
    short code appears in the class's LoDM AttTypeName list are considered
    LoDM-controlled.  Columns with an empty/absent Row 1 code (governance and
    metadata columns such as 'UAID_2', 'AssetName_2') are never added to the
    union set and therefore never blackened nor physically deleted.

    When *row1_codes* is None the legacy fuzzy description-based matching is
    used as a fallback.
    """
    class_col = next((c for c in columns if normalize_text(c) == "classification"), None)
    hier_col = next((c for c in columns if normalize_text(c) in ("assethierarchycategory", "hs2 class", "classcode")), None)
    au_columns = [c for i, c in enumerate(columns, start=1) if i >= au_start_idx]

    if logger:
        logger.info("    LoDM column mapping | hier_col=%r  class_col=%r  AW+ cols=%d",
                    hier_col, class_col, len(au_columns))

    class_cache: dict[str, set[str]] = {}
    row_sets: list[set[str]] = []
    union_set: set[str] = set()

    for _, row in mpdt_df.iterrows():
        code = ""
        # AssetHierarchyCategory holds the HS2 ClassCode which LoDM is keyed on.
        # Classification holds the Uniclass ID (e.g. 'Ss_45_30_05') which does NOT
        # match LoDM ClassCodes, so prioritise hier_col for LoDM attribute lookup.
        if hier_col and not _is_empty(row.get(hier_col)):
            code = str(row.get(hier_col)).strip()
        elif class_col and not _is_empty(row.get(class_col)):
            code = str(row.get(class_col)).strip()

        if code not in class_cache:
            if row1_codes is not None:
                # Normalised AttTypeName matching via Row 1 short codes.
                # _norm_att_code strips separators (space, _, -, :, .) and
                # lowercases so 'Com_Dscrptn' (template) matches 'Com:Dscrptn'
                # (LoDM).  Empty Row 1 codes (governance cols) never match.
                atttype_names = _lodm_atttype_names_for_class(lodm_df, code)
                matched = {
                    col for col in au_columns
                    if (r1 := _norm_att_code(row1_codes.get(col, ""))) and r1 in atttype_names
                }
                if logger:
                    logger.info(
                        "    LoDM mapping | ClassCode=%-20s | LoDM attrtypes=%3d | matched MPDT cols=%3d",
                        repr(code), len(atttype_names), len(matched),
                    )
            else:
                attr_norms = _lodm_attribute_norm_set_for_class(lodm_df, code)
                matched = _match_lodm_attrs_to_mpdt_columns(attr_norms, au_columns)
                if logger:
                    logger.info(
                        "    LoDM mapping | ClassCode=%-20s | LoDM attrs=%3d | matched MPDT cols=%3d",
                        repr(code), len(attr_norms), len(matched),
                    )
            class_cache[code] = matched
            if logger:
                for col in sorted(matched, key=lambda c: columns.index(c) if c in columns else 9999):
                    logger.debug("      [MATCH] %s", col.strip())
                unmatched_au = [c for c in au_columns if c not in matched]
                if unmatched_au and logger.isEnabledFor(logging.DEBUG):
                    for col in unmatched_au:
                        logger.debug("      [skip ] %s", col.strip())
        allowed = class_cache[code]
        row_sets.append(set(allowed))
        union_set.update(allowed)

    if logger:
        from openpyxl.utils import get_column_letter
        deleted_cols = [c for i, c in enumerate(columns, start=1) if i >= au_start_idx and c not in union_set]
        kept_cols = [c for i, c in enumerate(columns, start=1) if i >= au_start_idx and c in union_set]
        logger.info(
            "    LoDM union summary | kept=%d  deleted=%d (of %d AW+ cols)",
            len(kept_cols), len(deleted_cols), len(au_columns),
        )
        if deleted_cols:
            logger.info("    Columns to DELETE (not in any classcode LoDM):")
            for i, col in enumerate(deleted_cols):
                orig_idx = columns.index(col) + 1 if col in columns else "?"
                logger.info("      [DEL] col%-4s %s", get_column_letter(orig_idx) if isinstance(orig_idx, int) else orig_idx, col.strip())
        if kept_cols:
            logger.info("    Columns to KEEP (in LoDM for at least one classcode):")
            for col in kept_cols:
                orig_idx = columns.index(col) + 1 if col in columns else "?"
                logger.info("      [KEEP] col%-4s %s", get_column_letter(orig_idx) if isinstance(orig_idx, int) else orig_idx, col.strip())

    return row_sets, union_set




def _copy_row_format(ws, source_row: int, target_row: int, max_col: int) -> None:
    """Copy template row formatting to a generated data row without copying values."""
    if target_row == source_row:
        return
    ws.row_dimensions[target_row].height = ws.row_dimensions[source_row].height
    for col_idx in range(1, max_col + 1):
        src = ws.cell(row=source_row, column=col_idx)
        dst = ws.cell(row=target_row, column=col_idx)
        if src.has_style:
            dst._style = copy.copy(src._style)
        if src.number_format:
            dst.number_format = src.number_format
        if src.alignment:
            dst.alignment = copy.copy(src.alignment)
        if src.protection:
            dst.protection = copy.copy(src.protection)
        if src.hyperlink:
            dst._hyperlink = copy.copy(src.hyperlink)
        if src.comment:
            dst.comment = copy.copy(src.comment)


def _snapshot_row_fills(ws, template_row: int, max_col: int) -> dict[int, Any]:
    """Capture template fills before any checkerboard black fills are applied."""
    return {
        col_idx: copy.copy(ws.cell(row=template_row, column=col_idx).fill)
        for col_idx in range(1, max_col + 1)
    }


def _blank_fill_from_snapshot(ws, template_fills: dict[int, Any], target_row: int, col_idx: int) -> None:
    """Reset a data cell fill to the original template-row fill.

    Do not read from the live template row during checkerboard processing: row 3
    is also a generated data row and may already have been black-filled.
    """
    fill = template_fills.get(col_idx)
    if fill is not None:
        ws.cell(row=target_row, column=col_idx).fill = copy.copy(fill)


def _delete_columns_and_return_kept(
    ws,
    columns: list[str],
    delete_idxs: list[int],
) -> tuple[list[str], dict[str, int]]:
    """Delete columns from the worksheet and return remaining headers/1-based indexes.

    The returned list preserves duplicate pandas header names in order.  The map is
    keyed by header value, so duplicate blank/Unnamed headers will point to the last
    remaining occurrence; this is fine for the LoDM attribute headers, which are the
    only columns used by the applicability code.
    """
    delete_set = set(delete_idxs)
    kept_columns = [col for idx, col in enumerate(columns, start=1) if idx not in delete_set]
    for idx in sorted(delete_set, reverse=True):
        ws.delete_cols(idx, 1)
    kept_index_by_col = {col: idx for idx, col in enumerate(kept_columns, start=1)}
    return kept_columns, kept_index_by_col


# ---------------------------------------------------------------------------
# Pass 3: applicability (black fill for inapplicable cells)
# ---------------------------------------------------------------------------

_BLACK_FILL = PatternFill(fill_type="solid", fgColor="000000") if PatternFill else None


def build_att_map(wb_template) -> dict[str, str]:
    """Build {header_col: AttTst} from att_matrix sheet (col C → col A)."""
    att_map: dict[str, str] = {}
    if "att_matrix" not in wb_template.sheetnames:
        return att_map
    ws = wb_template["att_matrix"]
    for row_idx in range(1, ws.max_row + 1):
        key = ws.cell(row=row_idx, column=3).value
        val = ws.cell(row=row_idx, column=1).value
        if key and val:
            att_map[str(key).strip()] = str(val).strip()
    return att_map


def apply_applicability(
    ws,
    data_rows: list[dict[str, Any]],
    columns: list[str],
    row_allowed_cols: list[set[str]],
    start_row: int = 3,
    union_allowed_cols: "set[str] | None" = None,
    template_fills: dict[int, Any] | None = None,
) -> None:
    """Mark inapplicable AW+ cells with black fill based on class applicability.

    This must be called after any physical column deletion so the black cells are
    applied to the final worksheet coordinates.  Calling it before deleting AW+
    columns causes black fills to shift left into AU/AV and governance columns,
    which breaks the expected MPDT checkerboard pattern.
    """
    aw_start_idx = column_index_from_string("AW") if column_index_from_string else 49
    if aw_start_idx > len(columns):
        return

    for row_offset, _row_data in enumerate(data_rows):
        excel_row = start_row + row_offset
        allowed = row_allowed_cols[row_offset] if row_offset < len(row_allowed_cols) else set()
        for col_idx, col in enumerate(columns, start=1):
            if col_idx < aw_start_idx:
                # Defensive reset: AU/AV and earlier columns must never inherit
                # black fills from template rows or from shifted deleted columns.
                if template_fills is not None:
                    _blank_fill_from_snapshot(ws, template_fills, excel_row, col_idx)
                continue
            is_class_attr = union_allowed_cols is None or col in union_allowed_cols
            if is_class_attr and col not in allowed:
                ws.cell(row=excel_row, column=col_idx).fill = _BLACK_FILL
            else:
                if template_fills is not None:
                    _blank_fill_from_snapshot(ws, template_fills, excel_row, col_idx)


# ---------------------------------------------------------------------------
# Per-UAID MPDT generation
# ---------------------------------------------------------------------------

def generate_single_mpdt(
    workspace: Path,
    cfg: dict,
    uaid2: str,
    output_dir: Path,
    mapping_dict: dict[str, str],
    columns: list[str],
    scope2: pd.DataFrame,
    scope3: pd.DataFrame,
    sf_l2: pd.DataFrame,
    sf_l3: pd.DataFrame,
    lodm_df: pd.DataFrame,
    pw_df: pd.DataFrame,
    control_df: pd.DataFrame,
    logger: logging.Logger,
    deliverable_file: str = "",
    l2_df: "pd.DataFrame | None" = None,
    row1_codes: "dict[str, str] | None" = None,
    indexes: "dict[str, Any] | None" = None,
) -> Path | None:
    """Generate a single MPDT file for a UAID_2."""
    if load_workbook is None:
        raise ImportError("openpyxl is required. pip install openpyxl")

    template_path = (workspace / cfg.get("paths", {}).get("mpdt_template", "Input/C2-MPDT-Template-Mapping.xlsm")).resolve()
    if not template_path.exists():
        raise FileNotFoundError(f"MPDT template not found: {template_path}")

    logger.info("  Building joins for %s ...", uaid2)
    indexes = indexes or {}
    uaid2_key = str(uaid2).strip().upper()
    s3_rows_indexed = indexes.get("scope3_by_uaid2", {}).get(uaid2_key, pd.DataFrame())
    s2_row_indexed = indexes.get("scope2_by_uaid2", {}).get(uaid2_key)
    sf2_row_indexed = indexes.get("sf_l2_by_uaid2", {}).get(uaid2_key)

    if s2_row_indexed is not None or sf2_row_indexed is not None or not s3_rows_indexed.empty:
        if s2_row_indexed is not None and sf2_row_indexed is not None:
            join1_row = pd.Series(
                list(s2_row_indexed.values) + list(sf2_row_indexed.values),
                index=[f"{c}_l3" for c in s2_row_indexed.index] + [f"{c}_sform2" for c in sf2_row_indexed.index],
                dtype=object,
            )
        elif s2_row_indexed is not None:
            join1_row = s2_row_indexed
        elif sf2_row_indexed is not None:
            join1_row = sf2_row_indexed
        else:
            join1_row = pd.Series(dtype=object)

        join2_row = pd.Series(dtype=object)
        if not s3_rows_indexed.empty:
            uaid3_col_s3_i = indexes.get("uaid3_col_s3")
            sf3_by_uaid3 = indexes.get("sf_l3_by_uaid3", {})
            first_s3 = s3_rows_indexed.iloc[0]
            if uaid3_col_s3_i:
                uaid3_val = str(first_s3.get(uaid3_col_s3_i, "")).strip().upper()
                sf3_row = sf3_by_uaid3.get(uaid3_val)
                if sf3_row is not None:
                    join2_row = pd.Series(
                        list(sf3_row.values) + list(first_s3.values),
                        index=[f"{c}_sform3" for c in sf3_row.index] + [f"{c}_l3" for c in first_s3.index],
                        dtype=object,
                    )
                else:
                    join2_row = pd.Series(first_s3.values, index=[f"{c}_l3" for c in first_s3.index], dtype=object)
            else:
                join2_row = pd.Series(first_s3.values, index=[f"{c}_l3" for c in first_s3.index], dtype=object)
    else:
        join1_row, join2_row = build_joins(scope2, scope3, sf_l2, sf_l3, uaid2, logger)

    # Resolve ACBOS-based name from L2 mapping — mirrors notebook naming logic:
    #   acbos_value = str(acbos_match.iloc[0]['ACBOS']).replace('-ACBOS', '')
    #   filename = f"MPDT_{acbos_value}.xlsm"
    _acbos_stem = ""
    _l2_row = indexes.get("l2_by_uaid2", {}).get(uaid2_key)
    _acol = indexes.get("l2_acbos_col")
    if _l2_row is not None and _acol:
        _val = str(_l2_row.get(_acol, "")).strip()
        if _val and _val.lower() != "nan":
            _acbos_stem = _val.replace("-ACBOS", "")
    elif l2_df is not None and not l2_df.empty:
        _u2col = next((c for c in l2_df.columns if normalize_text(c) in ("uaid 2", "uaid2", "uaid")), None)
        _acol = next((c for c in l2_df.columns if normalize_text(c) in ("acbos doc", "acbos")), None)
        if _u2col and _acol:
            _m = l2_df[(l2_df[_u2col].fillna("").str.strip() == uaid2.strip()) &
                       (l2_df[_acol].notna()) &
                       (l2_df[_acol].astype(str).str.strip() != "")]
            if not _m.empty:
                _acbos_stem = str(_m.iloc[0][_acol]).strip().replace("-ACBOS", "")

    # Deliverable name to use for mappings that request the output filename
    if deliverable_file:
        deliverable_name = Path(deliverable_file).stem
    elif _acbos_stem:
        deliverable_name = f"MPDT_{_acbos_stem}"
    else:
        deliverable_name = f"MPDT_{uaid2}"

    model_container_id = ""
    resolver = indexes.get("model_container_resolver") if indexes else None
    if resolver is not None:
        try:
            model_container_id = resolver.resolve(uaid2)
            if model_container_id:
                logger.info("  Model Container ID resolved for %s: %s", uaid2, model_container_id)
            else:
                logger.info("  Model Container ID not uniquely resolved for %s; leaving blank", uaid2)
        except Exception as exc:
            logger.warning("  Could not resolve Model Container ID for %s: %s", uaid2, exc)
            model_container_id = ""

    # -----------------------------------------------------------------------
    # Optional diagnostic CSV dumps. Disabled by default because writing many
    # small CSV files for every target adds noticeable overhead. Enable with
    # config: {"debug_mpdt_generation": true}.
    # -----------------------------------------------------------------------
    if bool(cfg.get("debug_mpdt_generation", False)):
        debug_dir = output_dir / f"debug_{uaid2}"
        debug_dir.mkdir(parents=True, exist_ok=True)

        pd.DataFrame([join1_row]).to_csv(debug_dir / "join1_row.csv", index=False)
        pd.DataFrame([join2_row]).to_csv(debug_dir / "join2_row.csv", index=False)

        if not s3_rows_indexed.empty:
            s3_rows_indexed.to_csv(debug_dir / "scope3_rows.csv", index=False)
            logger.info("  Debug: join1=%d cols, join2=%d cols, scope3_rows=%d -> %s",
                        len(join1_row.index), len(join2_row.index), len(s3_rows_indexed), debug_dir)

        eval_rows = []
        for col in columns:
            expr = _mapping_expr_for_column(mapping_dict, col)
            val = get_mapped_value(expr, join2_row, join1_row, deliverable_name) if expr else None
            eval_rows.append({"mpdt_column": col, "mapping_expr": expr, "resolved_value": val})
        pd.DataFrame(eval_rows).to_csv(debug_dir / "mapping_eval.csv", index=False)
        logger.info("  Debug CSVs written to %s", debug_dir)

    # Build data rows (one per L3 asset under this L2)
    # Each scope3 row produces its own MPDT row with its own column values.
    data_rows: list[dict[str, Any]] = []
    rows_for_uaid = s3_rows_indexed
    if rows_for_uaid.empty and not scope3.empty:
        _uaid2_norms = {normalize_text(x) for x in ("UAID_2", "uaid_2", "uaid2", "ParentUaid", "Parent_UAID")}
        uaid2_col_s3 = next((c for c in scope3.columns if normalize_text(c) in _uaid2_norms), None)
        if uaid2_col_s3:
            rows_for_uaid = scope3[scope3[uaid2_col_s3].fillna("").astype(str).str.strip().str.upper() == uaid2.upper()]

    if rows_for_uaid.empty:
        logger.warning("  No scope3 rows found for %s — skipping MPDT.", uaid2)
        return None

    uaid3_col_s3 = indexes.get("uaid3_col_s3") or _uaid_col(scope3, ["UAID_3", "Uaid_3", "Uaid", "UAID"])
    sf3_by_uaid3 = indexes.get("sf_l3_by_uaid3", {})
    uaid3_col_sf = indexes.get("uaid3_col_sf") or (_uaid_col(sf_l3, ["UAID_3", "Uaid_3", "Uaid", "UAID", "Asset_ID"]) if not sf_l3.empty else None)

    for _, s3_row in rows_for_uaid.iterrows():
        if uaid3_col_s3:
            uaid3_val = str(s3_row.get(uaid3_col_s3, "")).strip().upper()
        else:
            uaid3_val = ""
        sf3_row = sf3_by_uaid3.get(uaid3_val)
        if sf3_row is not None:
            combined_idx = [f"{c}_l3" for c in s3_row.index] + [f"{c}_sform3" for c in sf3_row.index]
            combined_vals = list(s3_row.values) + list(sf3_row.values)
            row_join2 = pd.Series(combined_vals, index=combined_idx, dtype=object)
        elif uaid3_val and uaid3_col_sf and not sf_l3.empty:
            sf_match = sf_l3[sf_l3[uaid3_col_sf].fillna("").astype(str).str.strip().str.upper() == uaid3_val]
            if not sf_match.empty:
                sf_row = sf_match.iloc[0]
                combined_idx = [f"{c}_l3" for c in s3_row.index] + [f"{c}_sform3" for c in sf_row.index]
                combined_vals = list(s3_row.values) + list(sf_row.values)
                row_join2 = pd.Series(combined_vals, index=combined_idx, dtype=object)
            else:
                row_join2 = pd.Series(s3_row.values, index=[f"{c}_l3" for c in s3_row.index], dtype=object)
        else:
            row_join2 = pd.Series(s3_row.values, index=[f"{c}_l3" for c in s3_row.index], dtype=object)

        data_rows.append(populate_row(columns, mapping_dict, row_join2, join1_row, uaid2, deliverable_name, model_container_id))

    if not data_rows:
        logger.warning("  No scope3 data rows could be built for %s — skipping MPDT.", uaid2)
        return None

    # ------------------------------------------------------------------
    # Round 2: build DataFrame, compute derived columns, convert back
    # ------------------------------------------------------------------
    mpdt_df = pd.DataFrame(data_rows, columns=columns)
    sw_model_part_id_value = _extract_sw_model_part_id_from_existing_mpdt(
        workspace, cfg, uaid2, pw_df, columns, logger
    )
    mpdt_df = _apply_round2(mpdt_df, columns, uaid2, deliverable_name, sw_model_part_id_value, lodm_df, control_df, pw_df)
    data_rows = [{str(k): v for k, v in rec.items()} for rec in mpdt_df.to_dict("records")]

    # Output filename — mirror mpdt_creation_v2.ipynb naming:
    #   1. deliverable_file (from PW) takes priority
    #   2. l2_df ACBOS_Doc / ACBOS column (strip trailing -ACBOS) — _acbos_stem computed above
    #   3. fall back to MPDT_{uaid2}
    if deliverable_file:
        stem = Path(deliverable_file).stem
        filename = f"{stem}.xlsm"
    elif _acbos_stem:
        filename = f"MPDT_{_acbos_stem}.xlsm"
    else:
        filename = f"MPDT_{uaid2}.xlsm"

    output_dir.mkdir(parents=True, exist_ok=True)
    output_file = get_available_path(output_dir / sanitize_filename(filename))

    # Copy template and populate
    shutil.copy(template_path, output_file)
    wb = load_workbook(str(output_file), keep_vba=True)
    ws = wb["MPDT Element of Asset"]

    # Clear existing data (rows 3+)
    clear_start = 3
    for r in range(clear_start, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            ws.cell(row=r, column=c).value = None

    # Write data.  Preserve the template's data-row styles for every generated row;
    # openpyxl creates rows beyond the template with default formatting otherwise.
    for row_offset, row_data in enumerate(data_rows):
        excel_row = clear_start + row_offset
        _copy_row_format(ws, clear_start, excel_row, ws.max_column)
        for col_idx, col in enumerate(columns, start=1):
            ws.cell(row=excel_row, column=col_idx).value = row_data.get(col)

    # AW(49)+ = attribute columns subject to blackening and removal.
    # AU(47)=complete, AV(48)=count are excluded.
    aw_start_idx = column_index_from_string("AW") if column_index_from_string else 49
    row_allowed_cols, union_allowed_cols = _build_row_allowed_columns(
        mpdt_df, columns, lodm_df, aw_start_idx, logger, row1_codes=row1_codes
    )

    # Remove every non-permanent AW+ attribute column that is not required by
    # any ClassCode present in this MPDT. Permanent MPDT metadata columns
    # (UAID_2, AssetName_2, HS2_Class, chainage/grid fields, etc.) are kept by
    # an explicit allow-list. This prevents template-only or obsolete attributes
    # such as Vlm, LwstLvl, NmnlSrfcAr, FFwk, Ftop, and EmbdddLngth from leaking
    # into output files merely because they are absent from the current LoDM.
    _all_lodm_att_norms: set[str] = indexes.get("all_lodm_att_norms", set())
    if not _all_lodm_att_norms and row1_codes and not lodm_df.empty:
        _att_col2 = next(
            (c for c in lodm_df.columns if normalize_text(c) == "atttypename"), None
        )
        if _att_col2:
            _all_lodm_att_norms = {
                _norm_att_code(v)
                for v in lodm_df[_att_col2].dropna()
                if str(v).strip()
            }

    delete_idxs = sorted(
        [
            idx for idx, col in enumerate(columns, start=1)
            if idx >= aw_start_idx
            and col not in union_allowed_cols
            and not _is_permanent_mpdt_attribute_column(
                col, row1_codes.get(col, "") if row1_codes else ""
            )
        ],
        reverse=True,
    )
    # Delete first, then apply the black checkerboard fill to the final column
    # coordinates.  Applying fills before deletion shifts black formatting into
    # AU/AV and retained governance columns.
    kept_columns, _kept_index_by_col = _delete_columns_and_return_kept(ws, columns, delete_idxs)

    # Snapshot row-3 fills before applying black fills.  Row 3 is also a data row,
    # so using it as a live style source during checkerboard processing would make
    # later rows inherit row 3's black cells.
    template_fills = _snapshot_row_fills(ws, clear_start, ws.max_column)

    # Apply black fill from AW onwards only (AU=complete, AV=count are left as-is).
    # Pass union_allowed_cols so governance columns (never in any LoDM) stay white.
    apply_applicability(
        ws,
        data_rows,
        kept_columns,
        row_allowed_cols,
        start_row=clear_start,
        union_allowed_cols=union_allowed_cols,
        template_fills=template_fills,
    )

    wb.save(str(output_file))
    logger.info("  MPDT written: %s (%d data rows)", output_file.name, len(data_rows))
    return output_file


# ---------------------------------------------------------------------------
# Batch generation
# ---------------------------------------------------------------------------

def generate_mpdt_batch(
    workspace: Path,
    cfg: dict,
    targets: list[dict],
    output_dir: Path,
    sources: dict,
    logger: logging.Logger,
) -> dict:
    """
    Generate MPDT files for a batch of targets.
    Returns {generated: [...], errors: [...]}.
    """
    mapping_dict = sources["mapping_dict"]
    columns = sources["mpdt_columns"]
    row1_codes = sources.get("row1_codes")
    scope2 = sources["scope2_df"]
    scope3 = sources["scope3_df"]
    sf_l2 = sources["sf_l2"]
    sf_l3 = sources["sf_l3"]
    lodm_df = sources["lodm_df"]
    control_df = sources.get("control_df", pd.DataFrame())
    pw_df = sources.get("pw_df", pd.DataFrame())
    midp_df = sources.get("midp_df", pd.DataFrame())
    l2_df = sources.get("l2_df", None)

    if not columns:
        raise RuntimeError("Could not determine MPDT columns from sample file.")

    # Pre-index large input tables once per batch. The previous implementation
    # repeatedly scanned Scope3 and SmartForms L3 for every target and every L3
    # row; with large SmartForms files this dominates runtime.
    uaid2_col_s2 = _uaid_col(scope2, ["UAID_2", "Uaid_2", "Uaid", "UAID"])
    uaid2_col_sf2 = _uaid_col(sf_l2, ["UAID_2", "Uaid_2", "Uaid", "UAID", "Asset_ID"])
    uaid2_col_s3 = _uaid_col(scope3, ["UAID_2", "Uaid_2", "ParentUaid", "Parent_UAID"])
    uaid3_col_s3 = _uaid_col(scope3, ["UAID_3", "Uaid_3", "Uaid", "UAID"])
    uaid3_col_sf = _uaid_col(sf_l3, ["UAID_3", "Uaid_3", "Uaid", "UAID", "Asset_ID"]) if not sf_l3.empty else None
    l2_uaid_col = next((c for c in l2_df.columns if normalize_text(c) in ("uaid 2", "uaid2", "uaid")), None) if l2_df is not None and not l2_df.empty else None
    l2_acbos_col = next((c for c in l2_df.columns if normalize_text(c) in ("acbos doc", "acbos")), None) if l2_df is not None and not l2_df.empty else None

    all_lodm_att_norms: set[str] = set()
    if row1_codes and not lodm_df.empty:
        att_col = next((c for c in lodm_df.columns if normalize_text(c) == "atttypename"), None)
        if att_col:
            all_lodm_att_norms = {_norm_att_code(v) for v in lodm_df[att_col].dropna() if str(v).strip()}

    model_container_resolver = build_model_container_resolver(midp_df, pw_df, logger)

    indexes = {
        "model_container_resolver": model_container_resolver,
        "scope2_by_uaid2": _first_row_by_upper_key(scope2, uaid2_col_s2),
        "sf_l2_by_uaid2": _first_row_by_upper_key(sf_l2, uaid2_col_sf2),
        "scope3_by_uaid2": _group_df_by_upper_key(scope3, uaid2_col_s3),
        "sf_l3_by_uaid3": _first_row_by_upper_key(sf_l3, uaid3_col_sf),
        "l2_by_uaid2": _first_row_by_upper_key(l2_df, l2_uaid_col) if l2_df is not None else {},
        "l2_acbos_col": l2_acbos_col,
        "uaid3_col_s3": uaid3_col_s3,
        "uaid3_col_sf": uaid3_col_sf,
        "all_lodm_att_norms": all_lodm_att_norms,
    }
    logger.info(
        "Indexed sources: scope3 parents=%d, SmartForms L3 assets=%d, L2 rows=%d",
        len(indexes["scope3_by_uaid2"]), len(indexes["sf_l3_by_uaid3"]), len(indexes["l2_by_uaid2"]),
    )

    mpdt_dir = output_dir / "MPDT"
    mpdt_dir.mkdir(parents=True, exist_ok=True)

    generated, errors = [], []
    for target in targets:
        uaid2 = target["uaid"]
        deliverable_file = target.get("file", "")
        try:
            out = generate_single_mpdt(
                workspace, cfg, uaid2, mpdt_dir, mapping_dict, columns,
                scope2, scope3, sf_l2, sf_l3, lodm_df, pw_df, control_df, logger, deliverable_file,
                l2_df=l2_df,
                row1_codes=row1_codes,
                indexes=indexes,
            )
            if out:
                generated.append({"uaid": uaid2, "file": str(out)})
        except Exception as exc:
            logger.error("MPDT generation failed for %s: %s", uaid2, exc, exc_info=True)
            errors.append({"uaid": uaid2, "error": str(exc)})

    return {"generated": generated, "errors": errors}


# ---------------------------------------------------------------------------
# Standalone entry point
# ---------------------------------------------------------------------------

def main() -> None:
    parser = argparse.ArgumentParser(description="Generate MPDT files")
    parser.add_argument("--workspace", default=None)
    parser.add_argument("--target-uaid2", nargs="+", required=True)
    args = parser.parse_args()

    workspace = resolve_workspace(args.workspace)
    cfg = load_config(workspace)
    logger = setup_logger(workspace, "mpdt_generator", cfg.get("log_level", "INFO"))

    logger.info("=== MPDT Generator ===")

    from data_loader.local_loader import load_all_sources
    sources = load_all_sources(workspace, cfg, logger)

    uaids = []
    for v in args.target_uaid2:
        uaids.extend(u.strip() for u in v.split(",") if u.strip())

    targets = [{"uaid": u, "file": ""} for u in uaids]
    output_dir = timestamped_dir(workspace, "Output")

    result = generate_mpdt_batch(workspace, cfg, targets, output_dir, sources, logger)
    write_json(workspace / "Output" / "mpdt_result.json", result)
    logger.info("Generated: %d, Errors: %d", len(result["generated"]), len(result["errors"]))


if __name__ == "__main__":
    main()
