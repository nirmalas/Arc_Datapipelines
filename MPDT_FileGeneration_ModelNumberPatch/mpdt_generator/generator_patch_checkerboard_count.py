# generator_patch_checkerboard_count.py
import pandas as pd
from typing import List, Set

def lodm_attr_count_for_class(lodm_df: pd.DataFrame, class_code: str) -> int:
    if lodm_df is None or lodm_df.empty or not class_code:
        return 0
    cc_col = next((c for c in lodm_df.columns if str(c).strip().lower() == "classcode"), None)
    att_col = next((c for c in lodm_df.columns if str(c).strip().lower() == "atttypename"), None)
    if not cc_col or not att_col:
        return 0
    # Normalized lookup
    cc_variants = {str(class_code).strip().lower(), str(class_code).strip().replace("_","-").lower(), str(class_code).strip().replace("-","_").lower()}
    rows = lodm_df[lodm_df[cc_col].fillna("").astype(str).str.strip().str.lower().isin(cc_variants)]
    if rows.empty:
        return 0
    # Normalized unique AttTypeName
    _RE = re.compile(r'[\s_\-\.:]+')
    def _norm(s: str) -> str: return _RE.sub('', str(s or "")).lower()
    return len({_norm(v) for v in rows[att_col].dropna().astype(str) if str(v).strip()})

def _allowed_att_codes_for_class(lodm_df: pd.DataFrame, class_code: str) -> Set[str]:
    cc_col = next((c for c in lodm_df.columns if str(c).strip().lower() == "classcode"), None)
    att_col = next((c for c in lodm_df.columns if str(c).strip().lower() == "atttypename"), None)
    if not cc_col or not att_col:
        return set()
    variants = {str(class_code).strip().lower(), str(class_code).strip().replace("_","-").lower(), str(class_code).strip().replace("-","_").lower()}
    rows = lodm_df[lodm_df[cc_col].fillna("").astype(str).str.strip().str.lower().isin(variants)]
    _RE = re.compile(r'[\s_\-\.:]+')
    return {_RE.sub('', str(v)).lower() for v in rows[att_col].dropna().astype(str) if str(v).strip()}

def apply_checkerboard(ws, data_row_idxs: List[int], start_col_idx: int, row1_headers: List[str], class_codes: List[str], lodm_df: pd.DataFrame, first15_permanent: int = 15):
    """
    For each data row, blacken AU+ cells whose Row1 code is not allowed by the row's LoDM class.
    First 15 AU columns are always white (permanent).
    """
    from openpyxl.styles import PatternFill
    black_fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")

    for ridx, class_code in zip(data_row_idxs, class_codes):
        allowed = _allowed_att_codes_for_class(lodm_df, class_code)
        for offs, code in enumerate(row1_headers, start=0):
            col = start_col_idx + offs
            # Skip first 15 permanent AU columns
            if offs < first15_permanent:
                continue
            norm_code = re.sub(r'[\s_\-\.:]+', '', str(code or "")).lower()
            if norm_code and norm_code not in allowed:
                ws.cell(row=ridx, column=col).fill = black_fill

def write_count_for_row(ws, row_idx: int, count_col_idx: int, lodm_df: pd.DataFrame, class_code: str):
    cnt = lodm_attr_count_for_class(lodm_df, class_code)
    ws.cell(row=row_idx, column=count_col_idx, value=cnt + 10)