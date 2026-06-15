#!/usr/bin/env python3
"""Convert SmartForms fallback Excel to cached wide parquet files.

This script:
- Loads `Input/DB_Cache/mpdt_mapping_cache.json` to find attribute names referenced by MPDT mapping.
- Reads the SmartForms fallback Excel but only the key columns (`Uaid`, `AttributeTypeId`, `AttributeValue`).
- Filters attribute rows to those present in the mapping (case-insensitive).
- Pivots the filtered rows to a wide table (one row per UAID, columns=attributes).
- Writes caches to `Input/DB_Cache/SmartForms_L2(.parquet/.xlsx)` and `SmartForms_L3`.

Usage:
  python scripts/convert_smartforms_to_parquet.py
"""
from __future__ import annotations

import json
import logging
from pathlib import Path
import sys

import pandas as pd

# Ensure workspace root is on sys.path so `utils` imports resolve when run as a script
workspace = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(workspace))

from utils.common import resolve_workspace, load_config, setup_logger, write_table_cache


def main():
    workspace = resolve_workspace(None)
    cfg = load_config(workspace)
    logger = setup_logger(workspace, "convert_smartforms", cfg.get("log_level", "INFO"))

    cache_dir = workspace / cfg.get("paths", {}).get("db_cache_dir", "Input/DB_Cache")
    mapping_cache = cache_dir / "mpdt_mapping_cache.json"
    if not mapping_cache.exists():
        logger.error("MPDT mapping cache not found: %s", mapping_cache)
        raise SystemExit(1)

    payload = json.loads(mapping_cache.read_text(encoding="utf-8"))
    mapping = payload.get("mapping", {}) if isinstance(payload, dict) else {}
    # extract attribute names referenced in join2[...] expressions
    import re

    attr_names = set()
    for expr in mapping.values():
        if not isinstance(expr, str):
            continue
        for m in re.finditer(r"join[12]\s*\[\s*['\"]?([^'\"]+)['\"]?\s*\]", expr, re.IGNORECASE):
            attr_names.add(m.group(1).strip())

    if not attr_names:
        logger.warning("No attribute names discovered in mapping; falling back to default minimal set.")
        attr_names = {"UAID_3", "UAID_2", "Asset_ID"}

    logger.info("Discovered %d mapping attribute names to keep", len(attr_names))

    sf_file = workspace / cfg.get("paths", {}).get("smartforms_fallback", "Input/SmartForms_RAW_MPDT_L2&L3.xlsx")
    if not sf_file.exists():
        logger.error("SmartForms fallback file not found: %s", sf_file)
        raise SystemExit(1)

    xl = pd.ExcelFile(sf_file, engine="openpyxl")
    l2_sheet = next((s for s in xl.sheet_names if "l2" in s.lower()), xl.sheet_names[0])
    l3_sheet = next((s for s in xl.sheet_names if "l3" in s.lower()), xl.sheet_names[-1])

    key_cols = ["Uaid", "AttributeTypeId", "AttributeValue"]

    def _process(sheet_name, out_name):
        logger.info("Processing sheet '%s' -> cache '%s'", sheet_name, out_name)
        try:
            from openpyxl import load_workbook

            wb = load_workbook(filename=str(sf_file), read_only=True, data_only=True)
            ws = wb[sheet_name]
            rows = ws.iter_rows(values_only=True)
            header = next(rows)
            # map header names to indices
            hdr_map = {str(h).strip().lower(): i for i, h in enumerate(header) if h is not None}
            # possible header keys
            uaid_keys = ['uaid', 'uaid_3', 'uaid_2', 'uaid1', 'uaid2', 'uaid3', 'asset_id']
            attr_id_keys = ['attributetypeid']
            attr_code_keys = ['attrtypecode', 'attrtypecode']
            attr_display_keys = ['attrtypedisplayname', 'attrtypedescription', 'attrtypetext']
            attr_val_keys = ['attributevalue', 'attrvalue', 'attribute_value']

            def _find_index(candidates):
                for c in candidates:
                    if c in hdr_map:
                        return hdr_map[c]
                return None

            uaid_idx = _find_index(uaid_keys)
            attr_id_idx = _find_index(attr_id_keys)
            attr_code_idx = _find_index(attr_code_keys)
            attr_display_idx = _find_index(attr_display_keys)
            attr_val_idx = _find_index(attr_val_keys)

            if uaid_idx is None or attr_val_idx is None or (attr_id_idx is None and attr_code_idx is None and attr_display_idx is None):
                logger.warning('Could not find required key columns in sheet %s; falling back to pandas read', sheet_name)
                # fallback to pandas if indices not found
                df = pd.read_excel(sf_file, sheet_name=sheet_name, usecols=key_cols, dtype=str, engine="openpyxl")
            else:
                # First pass: gather attribute type counts to discover available attr codes/display names
                from collections import Counter

                attr_counter = Counter()
                sample_map = {}
                total = 0
                for r in rows:
                    total += 1
                    try:
                        attr_code = r[attr_code_idx] if attr_code_idx is not None else None
                        attr_disp = r[attr_display_idx] if attr_display_idx is not None else None
                        attr_id = r[attr_id_idx] if attr_id_idx is not None else None
                    except Exception:
                        continue
                    token = None
                    if attr_code and str(attr_code).strip():
                        token = str(attr_code).strip()
                    elif attr_disp and str(attr_disp).strip():
                        # display may be "Code - Description" - keep as-is
                        token = str(attr_disp).strip()
                    elif attr_id and str(attr_id).strip():
                        token = str(attr_id).strip()
                    if token:
                        attr_counter[token] += 1
                        sample_map[token] = (attr_code, attr_disp, attr_id)

                logger.info('Found %d distinct attribute types (scanned %d rows)', len(attr_counter), total)

                # Match mapping names to available tokens (case-insensitive substring heuristics)
                keep = {n.strip().lower() for n in attr_names}
                matched_tokens = set()
                for token in attr_counter.keys():
                    t_l = token.lower()
                    for m in keep:
                        if m in t_l or t_l in m:
                            matched_tokens.add(token)
                if not matched_tokens:
                    # as fallback, take top N frequent attribute types
                    TOP_N = 200
                    matched_tokens = set([t for t, _ in attr_counter.most_common(TOP_N)])
                    logger.info('No direct matches to mapping names; falling back to top-%d attribute types', TOP_N)

                # Second pass: re-open worksheet and collect rows for matched tokens
                wb2 = load_workbook(filename=str(sf_file), read_only=True, data_only=True)
                ws2 = wb2[sheet_name]
                rows2 = ws2.iter_rows(values_only=True)
                header2 = next(rows2)
                collected = []
                for r in rows2:
                    try:
                        uaid = r[uaid_idx] if uaid_idx is not None else None
                        val = r[attr_val_idx] if attr_val_idx is not None else None
                        attr_code = r[attr_code_idx] if attr_code_idx is not None else None
                        attr_disp = r[attr_display_idx] if attr_display_idx is not None else None
                        attr_id = r[attr_id_idx] if attr_id_idx is not None else None
                    except Exception:
                        continue
                    token = None
                    if attr_code and str(attr_code).strip():
                        token = str(attr_code).strip()
                    elif attr_disp and str(attr_disp).strip():
                        token = str(attr_disp).strip()
                    elif attr_id and str(attr_id).strip():
                        token = str(attr_id).strip()
                    if token and token in matched_tokens:
                        collected.append((str(uaid).strip() if uaid is not None else '', token, '' if val is None else str(val)))
                df = pd.DataFrame(collected, columns=['Uaid', 'AttributeTypeId', 'AttributeValue'])

        except Exception as exc:
            logger.warning('Streaming read failed for %s (%s); falling back to pandas', sheet_name, exc)
            df = pd.read_excel(sf_file, sheet_name=sheet_name, usecols=key_cols, dtype=str, engine="openpyxl")

        if df is None or df.empty:
            logger.warning("No attribute rows matched mapping attributes in sheet %s", sheet_name)
            write_table_cache(pd.DataFrame(), cache_dir / out_name, logger)
            return

        df['Uaid'] = df['Uaid'].astype(str).str.strip()
        df['AttributeValue'] = df['AttributeValue'].astype(str)

        # pivot
        try:
            pivoted = df.pivot_table(index='Uaid', columns='AttributeTypeId', values='AttributeValue', aggfunc='first')
            pivoted = pivoted.reset_index()
            # write cache
            write_table_cache(pivoted, cache_dir / out_name, logger)
            logger.info("Wrote pivoted cache %s — rows %d cols %d", out_name, len(pivoted), len(pivoted.columns))
        except Exception as exc:
            logger.error("Pivot failed for %s: %s", sheet_name, exc)

    _process(l2_sheet, 'SmartForms_L2')
    _process(l3_sheet, 'SmartForms_L3')


if __name__ == '__main__':
    main()
